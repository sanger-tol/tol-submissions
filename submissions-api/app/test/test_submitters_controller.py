# SPDX-FileCopyrightText: 2021 Genome Research Ltd.
#
# SPDX-License-Identifier: MIT

from __future__ import absolute_import

from test import BaseTestCase
from main.model import db, SubmissionsManifest, SubmissionsSpecimen, \
    SubmissionsSample
import main.excel_utils as excel_utils
import os
import responses
import tempfile
import filecmp
import datetime
from openpyxl import load_workbook
from unittest.mock import patch


class TestSubmittersController(BaseTestCase):

    def test_submit_manifest_json(self):
        # No authorisation token given
        body = []
        response = self.client.open(
            '/api/v1/manifests',
            method='POST',
            json=body)
        self.assert401(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Invalid authorisation token given
        body = []
        response = self.client.open(
            '/api/v1/manifests',
            method='POST',
            headers={"api-key": "12345678"},
            json=body)
        self.assert401(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Body empty
        body = {}
        response = self.client.open(
            '/api/v1/manifests',
            method='POST',
            headers={"api-key": self.user3.api_key},
            json=body)
        self.assert400(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        body = {'samples': [
                    {'row': 1,
                     'SPECIMEN_ID': 'SAN1234567',
                     'TAXON_ID': 6344,
                     'SCIENTIFIC_NAME': 'Arenicola marina',
                     'GENUS': 'Arenicola',
                     'FAMILY': 'Arenicolidae',
                     'ORDER_OR_GROUP': 'Scolecida',
                     'COMMON_NAME': 'lugworm',
                     'LIFESTAGE': 'ADULT',
                     'SEX': 'FEMALE',
                     'ORGANISM_PART': 'MUSCLE',
                     'GAL': 'SANGER INSTITUTE',
                     'GAL_SAMPLE_ID': 'SAN0000100',
                     'COLLECTED_BY': 'ALEX COLLECTOR',
                     'COLLECTOR_AFFILIATION': 'THE COLLECTOR INSTITUTE',
                     'DATE_OF_COLLECTION': '2020-09-01',
                     'COLLECTION_LOCATION': 'UNITED KINGDOM | DARK FOREST',
                     'DECIMAL_LATITUDE': '+50.12345678',
                     'DECIMAL_LONGITUDE': 'NOT_PROVIDED',
                     'HABITAT': 'Woodland',
                     'IDENTIFIED_BY': 'JO IDENTIFIER',
                     'IDENTIFIER_AFFILIATION': 'THE IDENTIFIER INSTITUTE',
                     'VOUCHER_ID': 'voucher1',
                     'OTHER_INFORMATION': 'Interesting',
                     'ELEVATION': '',  # Expect this to be set to None
                     'SERIES': '1',
                     'RACK_OR_PLATE_ID': 'RR12345678',
                     'TUBE_OR_WELL_ID': 'TU12345678',
                     'TAXON_REMARKS': 'nice',
                     'INFRASPECIFIC_EPITHET': 'ie1',
                     'COLLECTOR_SAMPLE_ID': 'mysample1',
                     'GRID_REFERENCE': 'TL 1234 5678',
                     'TIME_OF_COLLECTION': '12:00',
                     'DESCRIPTION_OF_COLLECTION_METHOD': 'picked it',
                     'DIFFICULT_OR_HIGH_PRIORITY_SAMPLE': 'Y',
                     'IDENTIFIED_HOW': 'looking',
                     'SPECIMEN_ID_RISK': 'Y',
                     'PRESERVED_BY': 'A Freezer',
                     'PRESERVER_AFFILIATION': 'cold storage ltd',
                     'PRESERVATION_APPROACH': 'Freezing',
                     'PRESERVATIVE_SOLUTION': 'Water',
                     'TIME_ELAPSED_FROM_COLLECTION_TO_PRESERVATION': '60',
                     'DATE_OF_PRESERVATION': '2021-04-04',
                     'SIZE_OF_TISSUE_IN_TUBE': 'VS',
                     'TISSUE_REMOVED_FOR_BARCODING': 'Y',
                     'PLATE_ID_FOR_BARCODING': 'plate1',
                     'TUBE_OR_WELL_ID_FOR_BARCODING': 'well1',
                     'TISSUE_FOR_BARCODING': 'LEG',
                     'BARCODE_PLATE_PRESERVATIVE': 'ethanol',
                     'PURPOSE_OF_SPECIMEN': 'REFERENCE_GENOME',
                     'HAZARD_GROUP': 'HG1',
                     'REGULATORY_COMPLIANCE': 'Y',
                     'ORIGINAL_COLLECTION_DATE': '2021-05-05',
                     'ORIGINAL_GEOGRAPHIC_LOCATION': 'United Kingdom | Light Forest',
                     'BARCODE_HUB': 'SANGER INSTITUTE',
                     'EXTRA_FIELD': 'extra1'}
                ]}
        # Not a submitter
        response = self.client.open(
            '/api/v1/manifests',
            method='POST',
            headers={"api-key": self.user1.api_key},
            json=body)
        self.assert403(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Correct, full JSON
        response = self.client.open(
            '/api/v1/manifests',
            method='POST',
            headers={"api-key": self.user3.api_key},
            json=body)
        self.assert200(response,
                       'Response body is : ' + response.data.decode('utf-8'))
        expected = {'manifestId': 1,
                    'submissionStatus': None,
                    'projectName': 'ToL',  # Default - not given in body
                    'stsManifestId': None,  # Default - not given in body
                    'samples': [
                        {'row': 1,
                         'SPECIMEN_ID': 'SAN1234567',
                         'TAXON_ID': 6344,
                         'SCIENTIFIC_NAME': 'Arenicola marina',
                         'GENUS': 'Arenicola',
                         'FAMILY': 'Arenicolidae',
                         'ORDER_OR_GROUP': 'Scolecida',
                         'COMMON_NAME': 'lugworm',
                         'LIFESTAGE': 'ADULT',
                         'SEX': 'FEMALE',
                         'ORGANISM_PART': 'MUSCLE',
                         'GAL': 'SANGER INSTITUTE',
                         'GAL_SAMPLE_ID': 'SAN0000100',
                         'COLLECTED_BY': 'ALEX COLLECTOR',
                         'COLLECTOR_AFFILIATION': 'THE COLLECTOR INSTITUTE',
                         'DATE_OF_COLLECTION': '2020-09-01',
                         'COLLECTION_LOCATION': 'UNITED KINGDOM | DARK FOREST',
                         'DECIMAL_LATITUDE': '+50.12345678',
                         'DECIMAL_LONGITUDE': 'NOT_PROVIDED',
                         'HABITAT': 'Woodland',
                         'IDENTIFIED_BY': 'JO IDENTIFIER',
                         'IDENTIFIER_AFFILIATION': 'THE IDENTIFIER INSTITUTE',
                         'VOUCHER_ID': 'voucher1',
                         'OTHER_INFORMATION': 'Interesting',
                         'ELEVATION': None,
                         'DEPTH': None,
                         'RELATIONSHIP': None,
                         'SYMBIONT': None,
                         'CULTURE_OR_STRAIN_ID': None,
                         'SERIES': '1',
                         'RACK_OR_PLATE_ID': 'RR12345678',
                         'TUBE_OR_WELL_ID': 'TU12345678',
                         'TAXON_REMARKS': 'nice',
                         'INFRASPECIFIC_EPITHET': 'ie1',
                         'COLLECTOR_SAMPLE_ID': 'mysample1',
                         'GRID_REFERENCE': 'TL 1234 5678',
                         'TIME_OF_COLLECTION': '12:00',
                         'DESCRIPTION_OF_COLLECTION_METHOD': 'picked it',
                         'DIFFICULT_OR_HIGH_PRIORITY_SAMPLE': 'Y',
                         'IDENTIFIED_HOW': 'looking',
                         'SPECIMEN_ID_RISK': 'Y',
                         'PRESERVED_BY': 'A Freezer',
                         'PRESERVER_AFFILIATION': 'cold storage ltd',
                         'PRESERVATION_APPROACH': 'Freezing',
                         'PRESERVATIVE_SOLUTION': 'Water',
                         'TIME_ELAPSED_FROM_COLLECTION_TO_PRESERVATION': '60',
                         'DATE_OF_PRESERVATION': '2021-04-04',
                         'SIZE_OF_TISSUE_IN_TUBE': 'VS',
                         'TISSUE_REMOVED_FOR_BARCODING': 'Y',
                         'PLATE_ID_FOR_BARCODING': 'plate1',
                         'TUBE_OR_WELL_ID_FOR_BARCODING': 'well1',
                         'TISSUE_FOR_BARCODING': 'LEG',
                         'BARCODE_PLATE_PRESERVATIVE': 'ethanol',
                         'PURPOSE_OF_SPECIMEN': 'REFERENCE_GENOME',
                         'HAZARD_GROUP': 'HG1',
                         'REGULATORY_COMPLIANCE': 'Y',
                         'ORIGINAL_COLLECTION_DATE': '2021-05-05',
                         'ORIGINAL_GEOGRAPHIC_LOCATION': 'United Kingdom | Light Forest',
                         'BARCODE_HUB': 'SANGER INSTITUTE',
                         'EXTRA_FIELD': 'extra1',
                         'tolId': None,
                         'biosampleAccession': None,
                         'sraAccession': None,
                         'submissionAccession': None,
                         'submissionError': None,
                         'sampleSameAs': None,
                         'sampleDerivedFrom': None,
                         'sampleSymbiontOf': None}
                    ]}
        print(response.json)
        self.assertEqual(expected, response.json)

    def test_get_manifest(self):

        body = {'samples': [
                    {'row': 1,
                     'SPECIMEN_ID': 'SAN1234567',
                     'TAXON_ID': 6344,
                     'SCIENTIFIC_NAME': 'Arenicola marina',
                     'GENUS': 'Arenicola',
                     'FAMILY': 'Arenicolidae',
                     'ORDER_OR_GROUP': 'Scolecida',
                     'COMMON_NAME': 'lugworm',
                     'LIFESTAGE': 'ADULT',
                     'SEX': 'FEMALE',
                     'ORGANISM_PART': 'MUSCLE',
                     'GAL': 'SANGER INSTITUTE',
                     'GAL_SAMPLE_ID': 'SAN0000100',
                     'COLLECTED_BY': 'ALEX COLLECTOR',
                     'COLLECTOR_AFFILIATION': 'THE COLLECTOR INSTITUTE',
                     'DATE_OF_COLLECTION': '2020-09-01',
                     'COLLECTION_LOCATION': 'UNITED KINGDOM | DARK FOREST',
                     'DECIMAL_LATITUDE': 'NOT_PROVIDED',
                     'DECIMAL_LONGITUDE': '-1.98765432',
                     'HABITAT': 'Woodland',
                     'IDENTIFIED_BY': 'JO IDENTIFIER',
                     'IDENTIFIER_AFFILIATION': 'THE IDENTIFIER INSTITUTE',
                     'VOUCHER_ID': 'voucher1',
                     'ELEVATION': None,  # Test None is same as missing
                     'DEPTH': ''}  # Test "" is the same as missing
                ],
                'projectName': 'TestProj',
                'stsManifestId': '1234-4321'}

        # Submit the manifest
        response = self.client.open(
            '/api/v1/manifests',
            method='POST',
            headers={"api-key": self.user3.api_key},
            json=body)
        self.assert200(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # No authorisation token given
        body = []
        response = self.client.open(
            '/api/v1/manifests/1',
            method='GET',
            json=body)
        self.assert401(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Invalid authorisation token given
        body = []
        response = self.client.open(
            '/api/v1/manifests/1',
            method='GET',
            headers={"api-key": "12345678"},
            json=body)
        self.assert401(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Incorrect manifest ID
        body = {}
        response = self.client.open(
            '/api/v1/manifests/2',
            method='GET',
            headers={"api-key": self.user3.api_key},
            json=body)
        self.assert404(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Not a submitter
        response = self.client.open(
            '/api/v1/manifests/1',
            method='GET',
            headers={"api-key": self.user1.api_key},
            json=body)
        self.assert403(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Corect - should validate and return errors
        response = self.client.open(
            '/api/v1/manifests/1',
            method='GET',
            headers={"api-key": self.user3.api_key},
            json=body)
        self.assert200(response,
                       'Response body is : ' + response.data.decode('utf-8'))
        expected = {'manifestId': 1,
                    'submissionStatus': None,
                    'projectName': 'TestProj',
                    'stsManifestId': '1234-4321',
                    'samples': [{
                        'row': 1,
                        'SPECIMEN_ID': 'SAN1234567',
                        'TAXON_ID': 6344,
                        'SCIENTIFIC_NAME': 'Arenicola marina',
                        'GENUS': 'Arenicola',
                        'FAMILY': 'Arenicolidae',
                        'ORDER_OR_GROUP': 'Scolecida',
                        'COMMON_NAME': 'lugworm',
                        'LIFESTAGE': 'ADULT',
                        'SEX': 'FEMALE',
                        'ORGANISM_PART': 'MUSCLE',
                        'GAL': 'SANGER INSTITUTE',
                        'GAL_SAMPLE_ID': 'SAN0000100',
                        'COLLECTED_BY': 'ALEX COLLECTOR',
                        'COLLECTOR_AFFILIATION': 'THE COLLECTOR INSTITUTE',
                        'DATE_OF_COLLECTION': '2020-09-01',
                        'COLLECTION_LOCATION': 'UNITED KINGDOM | DARK FOREST',
                        'DECIMAL_LATITUDE': 'NOT_PROVIDED',
                        'DECIMAL_LONGITUDE': '-1.98765432',
                        'HABITAT': 'Woodland',
                        'IDENTIFIED_BY': 'JO IDENTIFIER',
                        'IDENTIFIER_AFFILIATION': 'THE IDENTIFIER INSTITUTE',
                        'VOUCHER_ID': 'voucher1',
                        'OTHER_INFORMATION': None,
                        'DEPTH': None,
                        'ELEVATION': None,
                        'RELATIONSHIP': None,
                        'SYMBIONT': None,
                        'CULTURE_OR_STRAIN_ID': None,
                        'SERIES': None,
                        'RACK_OR_PLATE_ID': None,
                        'TUBE_OR_WELL_ID': None,
                        'TAXON_REMARKS': None,
                        'INFRASPECIFIC_EPITHET': None,
                        'COLLECTOR_SAMPLE_ID': None,
                        'GRID_REFERENCE': None,
                        'TIME_OF_COLLECTION': None,
                        'DESCRIPTION_OF_COLLECTION_METHOD': None,
                        'DIFFICULT_OR_HIGH_PRIORITY_SAMPLE': None,
                        'IDENTIFIED_HOW': None,
                        'SPECIMEN_ID_RISK': None,
                        'PRESERVED_BY': None,
                        'PRESERVER_AFFILIATION': None,
                        'PRESERVATION_APPROACH': None,
                        'PRESERVATIVE_SOLUTION': None,
                        'TIME_ELAPSED_FROM_COLLECTION_TO_PRESERVATION': None,
                        'DATE_OF_PRESERVATION': None,
                        'SIZE_OF_TISSUE_IN_TUBE': None,
                        'TISSUE_REMOVED_FOR_BARCODING': None,
                        'PLATE_ID_FOR_BARCODING': None,
                        'TUBE_OR_WELL_ID_FOR_BARCODING': None,
                        'TISSUE_FOR_BARCODING': None,
                        'BARCODE_PLATE_PRESERVATIVE': None,
                        'PURPOSE_OF_SPECIMEN': None,
                        'HAZARD_GROUP': None,
                        'REGULATORY_COMPLIANCE': None,
                        'ORIGINAL_COLLECTION_DATE': None,
                        'ORIGINAL_GEOGRAPHIC_LOCATION': None,
                        'BARCODE_HUB': None,
                        'tolId': None,
                        'biosampleAccession': None,
                        'sraAccession': None,
                        'submissionAccession': None,
                        'submissionError': None,
                        'sampleSameAs': None,
                        'sampleDerivedFrom': None,
                        'sampleSymbiontOf': None}
                    ]}
        self.assertEqual(expected, response.json)

    def test_get_manifests(self):

        manifest1 = SubmissionsManifest()
        manifest1.sts_manifest_id = "123-456-789"
        manifest1.project_name = "WowProj"
        manifest1.submission_status = True
        manifest1.created_at = datetime.datetime(2020, 5, 17)
        manifest1.user = self.user1
        db.session.add(manifest1)
        manifest2 = SubmissionsManifest()
        manifest2.sts_manifest_id = "987-654-321"
        manifest2.project_name = "DudProj"
        manifest2.submission_status = False
        manifest2.created_at = datetime.datetime(2021, 6, 18)
        manifest2.user = self.user2
        sample1 = SubmissionsSample(collected_by="ALEX COLLECTOR",
                                    collection_location="UNITED KINGDOM | DARK FOREST",
                                    collector_affiliation="THE COLLECTOR INSTITUTE",
                                    common_name="lugworm",
                                    date_of_collection="2020-09-01",
                                    decimal_latitude="50.12345678",
                                    decimal_longitude="-1.98765432",
                                    depth="100",
                                    elevation="0",
                                    family="Arenicolidae",
                                    GAL="SANGER INSTITUTE",
                                    GAL_sample_id="SAN0000100",
                                    genus="Arenicola",
                                    habitat="Woodland",
                                    identified_by="JO IDENTIFIER",
                                    identifier_affiliation="THE IDENTIFIER INSTITUTE",
                                    lifestage="ADULT",
                                    organism_part="MUSCLE",
                                    order_or_group="Scolecida",
                                    relationship="child of SAMEA1234567",
                                    scientific_name="Arenicola marina",
                                    sex="FEMALE",
                                    specimen_id="SAN0000100",
                                    taxonomy_id=6344,
                                    voucher_id="voucher1",
                                    row=1)
        sample1.manifest = manifest1
        db.session.add(manifest2)
        db.session.commit()

        # No authorisation token given
        body = []
        response = self.client.open(
            '/api/v1/manifests',
            method='GET',
            json=body)
        self.assert401(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Invalid authorisation token given
        body = []
        response = self.client.open(
            '/api/v1/manifests',
            method='GET',
            headers={"api-key": "12345678"},
            json=body)
        self.assert401(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Not a submitter
        response = self.client.open(
            '/api/v1/manifests',
            method='GET',
            headers={"api-key": self.user1.api_key},
            json=body)
        self.assert403(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Correct - should validate and return errors
        response = self.client.open(
            '/api/v1/manifests',
            method='GET',
            headers={"api-key": self.user3.api_key},
            json=body)
        self.assert200(response,
                       'Response body is : ' + response.data.decode('utf-8'))
        expected = [{'manifestId': 2,
                     'submissionStatus': False,
                     'projectName': 'DudProj',
                     'stsManifestId': '987-654-321',
                     'createdAt': '2021-06-18T00:00:00Z',
                     'numberOfSamples': 0,
                     'user': {'email': 'test_user_admin@sanger.ac.uk',
                              'name': 'test_user_admin',
                              'organisation': 'Sanger Institute',
                              'roles': [{'role': 'admin'}]}},
                    {'manifestId': 1,
                     'submissionStatus': True,
                     'projectName': 'WowProj',
                     'stsManifestId': '123-456-789',
                     'createdAt': '2020-05-17T00:00:00Z',
                     'numberOfSamples': 1,
                     'user': {'email': 'test_user_requester@sanger.ac.uk',
                              'name': 'test_user_requester',
                              'organisation': 'Sanger Institute',
                              'roles': []}}
                    ]
        self.assertEqual(expected, response.json)

    @responses.activate
    @patch('main.manifest_utils.get_ncbi_data')
    def test_validate_manifest_json(self, get_ncbi_data):
        mock_response_from_ena = {"taxId": "6344",
                                  "scientificName": "Arenicola marina",
                                  "commonName": "lugworm",
                                  "formalName": "true",
                                  "rank": "species",
                                  "division": "INV",
                                  "lineage": "",
                                  "geneticCode": "1",
                                  "mitochondrialGeneticCode": "5",
                                  "submittable": "true"}
        responses.add(responses.GET, 'https://www.ebi.ac.uk/ena/taxonomy/rest/tax-id/6344',
                      json=mock_response_from_ena, status=200)
        mock_response_from_tolid = [{"taxonomyId": "6344",
                                     "scientificName": "Arenicola marina",
                                     "commonName": "lugworm",
                                     "family": "Arenicolidae",
                                     "genus": "Arenicola",
                                     "order": "None"}]
        responses.add(responses.GET, os.environ['TOLID_URL'] + '/species/6344',
                      json=mock_response_from_tolid, status=200)

        mock_response_from_tolid_specimen = [{
            "specimenId": "SAN1234567",
            "tolIds": [{
                "species": {
                    "commonName": "lugworm",
                    "currentHighestTolidNumber": 2,
                    "family": "Arenicolidae",
                    "genus": "Arenicola",
                    "kingdom": "Metazoa",
                    "order": "Capitellida",
                    "phylum": "Annelida",
                    "prefix": "wuAreMari",
                    "scientificName": "Arenicola marina",
                    "taxaClass": "Polychaeta",
                    "taxonomyId": 6344
                },
                "tolId": "wuAreMari1"
            }]
        }]
        responses.add(responses.GET, os.environ['TOLID_URL'] + '/specimens/SAN1234567',
                      json=mock_response_from_tolid_specimen, status=200)

        mock_response_from_sts = {}  # Only interested in status codes
        responses.add(responses.GET, os.environ['STS_URL'] + '/samples/detail',
                      json=mock_response_from_sts, status=400)

        get_ncbi_data.return_value = {6344: {
            'TaxId': '6344',
            'ScientificName': 'Arenicola marina',
            'OtherNames': {
                'Anamorph': [],
                'CommonName': ['rock worm'],
                'Misnomer': [],
                'Inpart': [],
                'GenbankAnamorph': [],
                'Misspelling': [],
                'Includes': [],
                'EquivalentName': [],
                'Name': [{
                    'ClassCDE': 'authority',
                    'DispName': 'Arenicola marina (Linnaeus, 1758)'
                }, {
                    'ClassCDE': 'authority',
                    'DispName': 'Lumbricus marinus Linnaeus, 1758'
                }],
                'Synonym': ['Lumbricus marinus'],
                'GenbankSynonym': [],
                'Teleomorph': [],
                'Acronym': [],
                'GenbankCommonName': 'lugworm'},
            'ParentTaxId': '6343',
            'Rank': 'species',
            'Division': 'Invertebrates',
            'GeneticCode': {'GCId': '1', 'GCName': 'Standard'},
            'MitoGeneticCode': {'MGCId': '5', 'MGCName': 'Invertebrate Mitochondrial'},
            'Lineage': 'cellular organisms; Eukaryota; Opisthokonta; Metazoa; Eumetazoa; Bilateria; Protostomia; Spiralia; Lophotrochozoa; Annelida; Polychaeta; Sedentaria; Scolecida; Arenicolidae; Arenicola',  # noqa
            'LineageEx': [
                {'TaxId': '131567', 'ScientificName': 'cellular organisms', 'Rank': 'no rank'},
                {'TaxId': '2759', 'ScientificName': 'Eukaryota', 'Rank': 'superkingdom'},
                {'TaxId': '33154', 'ScientificName': 'Opisthokonta', 'Rank': 'clade'},
                {'TaxId': '33208', 'ScientificName': 'Metazoa', 'Rank': 'kingdom'},
                {'TaxId': '6072', 'ScientificName': 'Eumetazoa', 'Rank': 'clade'},
                {'TaxId': '33213', 'ScientificName': 'Bilateria', 'Rank': 'clade'},
                {'TaxId': '33317', 'ScientificName': 'Protostomia', 'Rank': 'clade'},
                {'TaxId': '2697495', 'ScientificName': 'Spiralia', 'Rank': 'clade'},
                {'TaxId': '1206795', 'ScientificName': 'Lophotrochozoa', 'Rank': 'clade'},
                {'TaxId': '6340', 'ScientificName': 'Annelida', 'Rank': 'phylum'},
                {'TaxId': '6341', 'ScientificName': 'Polychaeta', 'Rank': 'class'},
                {'TaxId': '105389', 'ScientificName': 'Sedentaria', 'Rank': 'subclass'},
                {'TaxId': '105387', 'ScientificName': 'Scolecida', 'Rank': 'infraclass'},
                {'TaxId': '42115', 'ScientificName': 'Arenicolidae', 'Rank': 'family'},
                {'TaxId': '6343', 'ScientificName': 'Arenicola', 'Rank': 'genus'}],
            'CreateDate': '1995/02/27 09: 24: 00',
            'UpdateDate': '2020/11/03 16: 20: 42',
            'PubDate': '1996/01/18 00: 00: 00'}
        }

        body = {'samples': [
                    {'row': 1,
                     'SPECIMEN_ID': 'SAN1234567',
                     'TAXON_ID': 6344,
                     'SCIENTIFIC_NAME': 'Arenicola marina',
                     'GENUS': 'Arenicola',
                     'FAMILY': 'Arenicolidae',
                     'ORDER_OR_GROUP': 'Scolecida',
                     'COMMON_NAME': 'lugworm',
                     'LIFESTAGE': 'ADULT',
                     'SEX': 'FEMALE',
                     'ORGANISM_PART': 'MUSCLE',
                     'GAL': 'SANGER INSTITUTE',
                     'GAL_SAMPLE_ID': 'SAN0000100',
                     'COLLECTED_BY': 'ALEX COLLECTOR',
                     'COLLECTOR_AFFILIATION': 'THE COLLECTOR INSTITUTE',
                     'DATE_OF_COLLECTION': '2020-09-01',
                     'COLLECTION_LOCATION': 'UNITED KINGDOM | DARK FOREST',
                     'DECIMAL_LATITUDE': '+50.12345678',
                     'DECIMAL_LONGITUDE': '-1.98765432',
                     'HABITAT': 'Woodland',
                     'IDENTIFIED_BY': 'JO IDENTIFIER',
                     'IDENTIFIER_AFFILIATION': 'THE IDENTIFIER INSTITUTE',
                     'VOUCHER_ID': ''}
                ]}

        # Submit the manifest
        response = self.client.open(
            '/api/v1/manifests',
            method='POST',
            headers={"api-key": self.user3.api_key},
            json=body)
        self.assert200(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # No authorisation token given
        body = []
        response = self.client.open(
            '/api/v1/manifests/1/validate',
            method='GET',
            json=body)
        self.assert401(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Invalid authorisation token given
        body = []
        response = self.client.open(
            '/api/v1/manifests/1/validate',
            method='GET',
            headers={"api-key": "12345678"},
            json=body)
        self.assert401(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Incorrect manifest ID
        body = {}
        response = self.client.open(
            '/api/v1/manifests/2/validate',
            method='GET',
            headers={"api-key": self.user3.api_key},
            json=body)
        self.assert404(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Not a submitter
        response = self.client.open(
            '/api/v1/manifests/1/validate',
            method='GET',
            headers={"api-key": self.user1.api_key},
            json=body)
        self.assert403(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Corect - should validate and return errors
        response = self.client.open(
            '/api/v1/manifests/1/validate',
            method='GET',
            headers={"api-key": self.user3.api_key},
            json=body)
        self.assert200(response,
                       'Response body is : ' + response.data.decode('utf-8'))
        expected = {'manifestId': 1,
                    "number_of_errors": 1,
                    'validations': [
                        {"row": 1,
                         "results": [
                             {"field": "VOUCHER_ID",
                              "message": "Must not be empty",
                              'severity': 'ERROR'}
                         ]}
                    ]}
        self.assertEqual(expected, response.json)

    @responses.activate
    @patch('main.manifest_utils.get_ncbi_data')
    def test_submit_and_validate_manifest_json(self, get_ncbi_data):
        mock_response_from_ena = {"taxId": "6344",
                                  "scientificName": "Arenicola marina",
                                  "commonName": "lugworm",
                                  "formalName": "true",
                                  "rank": "species",
                                  "division": "INV",
                                  "lineage": "",
                                  "geneticCode": "1",
                                  "mitochondrialGeneticCode": "5",
                                  "submittable": "true"}
        responses.add(responses.GET, 'https://www.ebi.ac.uk/ena/taxonomy/rest/tax-id/6344',
                      json=mock_response_from_ena, status=200)
        mock_response_from_tolid = [{"taxonomyId": "6344",
                                     "scientificName": "Arenicola marina",
                                     "commonName": "lugworm",
                                     "family": "Arenicolidae",
                                     "genus": "Arenicola",
                                     "order": "None"}]
        responses.add(responses.GET, os.environ['TOLID_URL'] + '/species/6344',
                      json=mock_response_from_tolid, status=200)

        mock_response_from_tolid_specimen = [{
            "specimenId": "SAN1234567",
            "tolIds": [{
                "species": {
                    "commonName": "lugworm",
                    "currentHighestTolidNumber": 2,
                    "family": "Arenicolidae",
                    "genus": "Arenicola",
                    "kingdom": "Metazoa",
                    "order": "Capitellida",
                    "phylum": "Annelida",
                    "prefix": "wuAreMari",
                    "scientificName": "Arenicola marina",
                    "taxaClass": "Polychaeta",
                    "taxonomyId": 6344
                },
                "tolId": "wuAreMari1"
            }]
        }]
        responses.add(responses.GET, os.environ['TOLID_URL'] + '/specimens/SAN1234567',
                      json=mock_response_from_tolid_specimen, status=200)

        mock_response_from_sts = {}  # Only interested in status codes
        responses.add(responses.GET, os.environ['STS_URL'] + '/samples/detail',
                      json=mock_response_from_sts, status=400)

        get_ncbi_data.return_value = {6344: {
            'TaxId': '6344',
            'ScientificName': 'Arenicola marina',
            'OtherNames': {
                'Anamorph': [],
                'CommonName': ['rock worm'],
                'Misnomer': [],
                'Inpart': [],
                'GenbankAnamorph': [],
                'Misspelling': [],
                'Includes': [],
                'EquivalentName': [],
                'Name': [{
                    'ClassCDE': 'authority',
                    'DispName': 'Arenicola marina (Linnaeus, 1758)'
                }, {
                    'ClassCDE': 'authority',
                    'DispName': 'Lumbricus marinus Linnaeus, 1758'
                }],
                'Synonym': ['Lumbricus marinus'],
                'GenbankSynonym': [],
                'Teleomorph': [],
                'Acronym': [],
                'GenbankCommonName': 'lugworm'},
            'ParentTaxId': '6343',
            'Rank': 'species',
            'Division': 'Invertebrates',
            'GeneticCode': {'GCId': '1', 'GCName': 'Standard'},
            'MitoGeneticCode': {'MGCId': '5', 'MGCName': 'Invertebrate Mitochondrial'},
            'Lineage': 'cellular organisms; Eukaryota; Opisthokonta; Metazoa; Eumetazoa; Bilateria; Protostomia; Spiralia; Lophotrochozoa; Annelida; Polychaeta; Sedentaria; Scolecida; Arenicolidae; Arenicola',  # noqa
            'LineageEx': [
                {'TaxId': '131567', 'ScientificName': 'cellular organisms', 'Rank': 'no rank'},
                {'TaxId': '2759', 'ScientificName': 'Eukaryota', 'Rank': 'superkingdom'},
                {'TaxId': '33154', 'ScientificName': 'Opisthokonta', 'Rank': 'clade'},
                {'TaxId': '33208', 'ScientificName': 'Metazoa', 'Rank': 'kingdom'},
                {'TaxId': '6072', 'ScientificName': 'Eumetazoa', 'Rank': 'clade'},
                {'TaxId': '33213', 'ScientificName': 'Bilateria', 'Rank': 'clade'},
                {'TaxId': '33317', 'ScientificName': 'Protostomia', 'Rank': 'clade'},
                {'TaxId': '2697495', 'ScientificName': 'Spiralia', 'Rank': 'clade'},
                {'TaxId': '1206795', 'ScientificName': 'Lophotrochozoa', 'Rank': 'clade'},
                {'TaxId': '6340', 'ScientificName': 'Annelida', 'Rank': 'phylum'},
                {'TaxId': '6341', 'ScientificName': 'Polychaeta', 'Rank': 'class'},
                {'TaxId': '105389', 'ScientificName': 'Sedentaria', 'Rank': 'subclass'},
                {'TaxId': '105387', 'ScientificName': 'Scolecida', 'Rank': 'infraclass'},
                {'TaxId': '42115', 'ScientificName': 'Arenicolidae', 'Rank': 'family'},
                {'TaxId': '6343', 'ScientificName': 'Arenicola', 'Rank': 'genus'}],
            'CreateDate': '1995/02/27 09: 24: 00',
            'UpdateDate': '2020/11/03 16: 20: 42',
            'PubDate': '1996/01/18 00: 00: 00'}
        }

        # No authorisation token given
        body = []
        response = self.client.open(
            '/api/v1/manifests/validate',
            method='POST',
            json=body)
        self.assert401(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Invalid authorisation token given
        body = []
        response = self.client.open(
            '/api/v1/manifests/validate',
            method='POST',
            headers={"api-key": "12345678"},
            json=body)
        self.assert401(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Body empty
        body = {}
        response = self.client.open(
            '/api/v1/manifests/validate',
            method='POST',
            headers={"api-key": self.user3.api_key},
            json=body)
        self.assert400(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        body = {'samples': [
                    {'row': 1,
                     'SPECIMEN_ID': 'SAN1234567',
                     'TAXON_ID': 6344,
                     'SCIENTIFIC_NAME': 'Arenicola marina',
                     'GENUS': 'Arenicola',
                     'FAMILY': 'Arenicolidae',
                     'ORDER_OR_GROUP': 'Scolecida',
                     'COMMON_NAME': 'lugworm',
                     'LIFESTAGE': 'ADULT',
                     'SEX': 'FEMALE',
                     'ORGANISM_PART': 'MUSCLE',
                     'GAL': 'SANGER INSTITUTE',
                     'GAL_SAMPLE_ID': 'SAN0000100',
                     'COLLECTED_BY': 'ALEX COLLECTOR',
                     'COLLECTOR_AFFILIATION': 'THE COLLECTOR INSTITUTE',
                     'DATE_OF_COLLECTION': '2020-09-01',
                     'COLLECTION_LOCATION': 'UNITED KINGDOM | DARK FOREST',
                     'DECIMAL_LATITUDE': '+50.12345678',
                     'DECIMAL_LONGITUDE': '-1.98765432',
                     'HABITAT': 'Woodland',
                     'IDENTIFIED_BY': 'JO IDENTIFIER',
                     'IDENTIFIER_AFFILIATION': 'THE IDENTIFIER INSTITUTE',
                     'VOUCHER_ID': ''}
                ]}
        # Not a submitter
        response = self.client.open(
            '/api/v1/manifests/validate',
            method='POST',
            headers={"api-key": self.user1.api_key},
            json=body)
        self.assert403(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Correct, full JSON
        response = self.client.open(
            '/api/v1/manifests/validate',
            method='POST',
            headers={"api-key": self.user3.api_key},
            json=body)
        self.assert200(response,
                       'Response body is : ' + response.data.decode('utf-8'))
        expected = {'manifestId': 1,
                    "number_of_errors": 1,
                    'validations': [
                        {"row": 1,
                         "results": [
                             {"field": "VOUCHER_ID",
                              "message": "Must not be empty",
                              'severity': 'ERROR'}
                         ]}
                    ]}
        self.assertEqual(expected, response.json)

    @responses.activate
    def test_generate_ids(self):
        specimen = SubmissionsSpecimen()
        specimen.specimen_id = "SAN1234567"
        specimen.biosample_accession = "SAMEA12345678"
        db.session.add(specimen)
        db.session.commit()

        mock_response_from_ena = {"taxId": "6344",
                                  "scientificName": "Arenicola marina",
                                  "commonName": "lugworm",
                                  "formalName": "true",
                                  "rank": "species",
                                  "division": "INV",
                                  "lineage": "",
                                  "geneticCode": "1",
                                  "mitochondrialGeneticCode": "5",
                                  "submittable": "true"}
        responses.add(responses.GET, 'https://www.ebi.ac.uk/ena/taxonomy/rest/tax-id/6344',
                      json=mock_response_from_ena, status=200)
        mock_response_from_tolid = [{
            "species": {
                "commonName": "lugworm",
                "family": "Arenicolidae",
                "genus": "Arenicola",
                "kingdom": "Metazoa",
                "order": "Capitellida",
                "phylum": "Annelida",
                "prefix": "wuAreMari",
                "scientificName": "Arenicola marina",
                "taxaClass": "Polychaeta",
                "taxonomyId": 6344
            },
            "specimen": {
                "specimenId": "SAN1234567"
            },
            "tolId": "wuAreMari1"
        }]
        responses.add(responses.POST, os.environ['TOLID_URL'] + '/tol-ids',
                      json=mock_response_from_tolid, status=200)

        body = {'samples': [
                    {'row': 1,
                     'SPECIMEN_ID': 'SAN1234567',
                     'TAXON_ID': 6344,
                     'SCIENTIFIC_NAME': 'Arenicola marina',
                     'GENUS': 'Arenicola',
                     'FAMILY': 'Arenicolidae',
                     'ORDER_OR_GROUP': 'Scolecida',
                     'COMMON_NAME': 'lugworm',
                     'LIFESTAGE': 'ADULT',
                     'SEX': 'FEMALE',
                     'ORGANISM_PART': 'MUSCLE',
                     'GAL': 'SANGER INSTITUTE',
                     'GAL_SAMPLE_ID': 'SAN0000100',
                     'COLLECTED_BY': 'ALEX COLLECTOR',
                     'COLLECTOR_AFFILIATION': 'THE COLLECTOR INSTITUTE',
                     'DATE_OF_COLLECTION': '2020-09-01',
                     'COLLECTION_LOCATION': 'UNITED KINGDOM | DARK FOREST',
                     'DECIMAL_LATITUDE': '+50.12345678',
                     'DECIMAL_LONGITUDE': '-1.98765432',
                     'HABITAT': 'Woodland',
                     'IDENTIFIED_BY': 'JO IDENTIFIER',
                     'IDENTIFIER_AFFILIATION': 'THE IDENTIFIER INSTITUTE',
                     'VOUCHER_ID': 'voucher1'}
                ]}

        # Submit the manifest
        response = self.client.open(
            '/api/v1/manifests',
            method='POST',
            headers={"api-key": self.user3.api_key},
            json=body)
        self.assert200(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        manifest = db.session.query(SubmissionsManifest) \
            .filter(SubmissionsManifest.manifest_id == 1) \
            .one_or_none()
        mock_response_from_ena = '<?xml version="1.0" encoding="UTF-8"?><?xml-stylesheet type="text/xsl" href="receipt.xsl"?><RECEIPT receiptDate="2021-04-07T12:47:39.998+01:00" submissionFile="tmphz9luulrsubmission_3.xml" success="true"><SAMPLE accession="ERS6206028" alias="' + str(manifest.samples[0].sample_id) + '" status="PRIVATE"><EXT_ID accession="SAMEA8521239" type="biosample"/></SAMPLE><SUBMISSION accession="ERA3819349" alias="SUBMISSION-07-04-2021-12:47:36:825"/><ACTIONS>ADD</ACTIONS></RECEIPT>'  # noqa
        responses.add(responses.POST, os.environ['ENA_URL'] + '/ena/submit/drop-box/submit/',
                      body=mock_response_from_ena, status=200)

        # No authorisation token given
        body = []
        response = self.client.open(
            '/api/v1/manifests/1/generate',
            method='PATCH',
            json=body)
        self.assert401(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Invalid authorisation token given
        body = []
        response = self.client.open(
            '/api/v1/manifests/1/generate',
            method='PATCH',
            headers={"api-key": "12345678"},
            json=body)
        self.assert401(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Incorrect manifest ID
        body = {}
        response = self.client.open(
            '/api/v1/manifests/2/generate',
            method='PATCH',
            headers={"api-key": self.user3.api_key},
            json=body)
        self.assert404(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Not a submitter
        response = self.client.open(
            '/api/v1/manifests/1/generate',
            method='PATCH',
            headers={"api-key": self.user1.api_key},
            json=body)
        self.assert403(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Corect - should validate and return errors
        response = self.client.open(
            '/api/v1/manifests/1/generate',
            method='PATCH',
            headers={"api-key": self.user3.api_key},
            json=body)
        self.assert200(response,
                       'Response body is : ' + response.data.decode('utf-8'))
        expected = {'manifestId': 1,
                    'submissionStatus': True,
                    'projectName': 'ToL',
                    'stsManifestId': None,
                    'samples': [{
                        'row': 1,
                        'SPECIMEN_ID': 'SAN1234567',
                        'TAXON_ID': 6344,
                        'SCIENTIFIC_NAME': 'Arenicola marina',
                        'GENUS': 'Arenicola',
                        'FAMILY': 'Arenicolidae',
                        'ORDER_OR_GROUP': 'Scolecida',
                        'COMMON_NAME': 'lugworm',
                        'LIFESTAGE': 'ADULT',
                        'SEX': 'FEMALE',
                        'ORGANISM_PART': 'MUSCLE',
                        'GAL': 'SANGER INSTITUTE',
                        'GAL_SAMPLE_ID': 'SAN0000100',
                        'COLLECTED_BY': 'ALEX COLLECTOR',
                        'COLLECTOR_AFFILIATION': 'THE COLLECTOR INSTITUTE',
                        'DATE_OF_COLLECTION': '2020-09-01',
                        'COLLECTION_LOCATION': 'UNITED KINGDOM | DARK FOREST',
                        'DECIMAL_LATITUDE': '+50.12345678',
                        'DECIMAL_LONGITUDE': '-1.98765432',
                        'HABITAT': 'Woodland',
                        'IDENTIFIED_BY': 'JO IDENTIFIER',
                        'IDENTIFIER_AFFILIATION': 'THE IDENTIFIER INSTITUTE',
                        'VOUCHER_ID': 'voucher1',
                        'OTHER_INFORMATION': None,
                        'DEPTH': None,
                        'ELEVATION': None,
                        'RELATIONSHIP': None,
                        'SYMBIONT': None,
                        'CULTURE_OR_STRAIN_ID': None,
                        'SERIES': None,
                        'RACK_OR_PLATE_ID': None,
                        'TUBE_OR_WELL_ID': None,
                        'TAXON_REMARKS': None,
                        'INFRASPECIFIC_EPITHET': None,
                        'COLLECTOR_SAMPLE_ID': None,
                        'GRID_REFERENCE': None,
                        'TIME_OF_COLLECTION': None,
                        'DESCRIPTION_OF_COLLECTION_METHOD': None,
                        'DIFFICULT_OR_HIGH_PRIORITY_SAMPLE': None,
                        'IDENTIFIED_HOW': None,
                        'SPECIMEN_ID_RISK': None,
                        'PRESERVED_BY': None,
                        'PRESERVER_AFFILIATION': None,
                        'PRESERVATION_APPROACH': None,
                        'PRESERVATIVE_SOLUTION': None,
                        'TIME_ELAPSED_FROM_COLLECTION_TO_PRESERVATION': None,
                        'DATE_OF_PRESERVATION': None,
                        'SIZE_OF_TISSUE_IN_TUBE': None,
                        'TISSUE_REMOVED_FOR_BARCODING': None,
                        'PLATE_ID_FOR_BARCODING': None,
                        'TUBE_OR_WELL_ID_FOR_BARCODING': None,
                        'TISSUE_FOR_BARCODING': None,
                        'BARCODE_PLATE_PRESERVATIVE': None,
                        'PURPOSE_OF_SPECIMEN': None,
                        'HAZARD_GROUP': None,
                        'REGULATORY_COMPLIANCE': None,
                        'ORIGINAL_COLLECTION_DATE': None,
                        'ORIGINAL_GEOGRAPHIC_LOCATION': None,
                        'BARCODE_HUB': None,
                        'tolId': 'wuAreMari1',
                        'biosampleAccession': "SAMEA8521239",
                        'sraAccession': "ERS6206028",
                        'submissionAccession': "ERA3819349",
                        'submissionError': None,
                        'sampleSameAs': None,
                        'sampleDerivedFrom': 'SAMEA12345678',
                        'sampleSymbiontOf': None}
                    ]}
        self.assertEqual(expected, response.json)

    @responses.activate
    def test_submit_and_generate_manifest_json(self):
        specimen = SubmissionsSpecimen()
        specimen.specimen_id = "SAN1234567"
        specimen.biosample_accession = "SAMEA12345678"
        db.session.add(specimen)
        db.session.commit()

        mock_response_from_ena = {"taxId": "6344",
                                  "scientificName": "Arenicola marina",
                                  "commonName": "lugworm",
                                  "formalName": "true",
                                  "rank": "species",
                                  "division": "INV",
                                  "lineage": "",
                                  "geneticCode": "1",
                                  "mitochondrialGeneticCode": "5",
                                  "submittable": "true"}
        responses.add(responses.GET, 'https://www.ebi.ac.uk/ena/taxonomy/rest/tax-id/6344',
                      json=mock_response_from_ena, status=200)
        mock_response_from_tolid = [{
            "species": {
                "commonName": "lugworm",
                "family": "Arenicolidae",
                "genus": "Arenicola",
                "kingdom": "Metazoa",
                "order": "Capitellida",
                "phylum": "Annelida",
                "prefix": "wuAreMari",
                "scientificName": "Arenicola marina",
                "taxaClass": "Polychaeta",
                "taxonomyId": 6344
            },
            "specimen": {
                "specimenId": "SAN1234567"
            },
            "tolId": "wuAreMari1"
        }]
        responses.add(responses.POST, os.environ['TOLID_URL'] + '/tol-ids',
                      json=mock_response_from_tolid, status=200)

        mock_response_from_ena = '<?xml version="1.0" encoding="UTF-8"?><?xml-stylesheet type="text/xsl" href="receipt.xsl"?><RECEIPT receiptDate="2021-04-07T12:47:39.998+01:00" submissionFile="tmphz9luulrsubmission_3.xml" success="true"><SAMPLE accession="ERS6206028" alias="' + str(1) + '" status="PRIVATE"><EXT_ID accession="SAMEA8521239" type="biosample"/></SAMPLE><SUBMISSION accession="ERA3819349" alias="SUBMISSION-07-04-2021-12:47:36:825"/><ACTIONS>ADD</ACTIONS></RECEIPT>'  # noqa
        responses.add(responses.POST, os.environ['ENA_URL'] + '/ena/submit/drop-box/submit/',
                      body=mock_response_from_ena, status=200)

        body = {'samples': [
                    {'row': 1,
                     'SPECIMEN_ID': 'SAN1234567',
                     'TAXON_ID': 6344,
                     'SCIENTIFIC_NAME': 'Arenicola marina',
                     'GENUS': 'Arenicola',
                     'FAMILY': 'Arenicolidae',
                     'ORDER_OR_GROUP': 'Scolecida',
                     'COMMON_NAME': 'lugworm',
                     'LIFESTAGE': 'ADULT',
                     'SEX': 'FEMALE',
                     'ORGANISM_PART': 'MUSCLE',
                     'GAL': 'SANGER INSTITUTE',
                     'GAL_SAMPLE_ID': 'SAN0000100',
                     'COLLECTED_BY': 'ALEX COLLECTOR',
                     'COLLECTOR_AFFILIATION': 'THE COLLECTOR INSTITUTE',
                     'DATE_OF_COLLECTION': '2020-09-01',
                     'COLLECTION_LOCATION': 'UNITED KINGDOM | DARK FOREST',
                     'DECIMAL_LATITUDE': '+50.12345678',
                     'DECIMAL_LONGITUDE': '-1.98765432',
                     'HABITAT': 'Woodland',
                     'IDENTIFIED_BY': 'JO IDENTIFIER',
                     'IDENTIFIER_AFFILIATION': 'THE IDENTIFIER INSTITUTE',
                     'VOUCHER_ID': 'voucher1'}
                ]}

        # No authorisation token given
        response = self.client.open(
            '/api/v1/manifests/generate',
            method='POST',
            json=body)
        self.assert401(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Invalid authorisation token given
        response = self.client.open(
            '/api/v1/manifests/generate',
            method='POST',
            headers={"api-key": "12345678"},
            json=body)
        self.assert401(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Not a submitter
        response = self.client.open(
            '/api/v1/manifests/generate',
            method='POST',
            headers={"api-key": self.user1.api_key},
            json=body)
        self.assert403(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Corect - should validate and return errors
        response = self.client.open(
            '/api/v1/manifests/generate',
            method='POST',
            headers={"api-key": self.user3.api_key},
            json=body)
        self.assert200(response,
                       'Response body is : ' + response.data.decode('utf-8'))
        expected = {'manifestId': 1,
                    'submissionStatus': True,
                    'projectName': 'ToL',
                    'stsManifestId': None,
                    'samples': [{
                        'row': 1,
                        'SPECIMEN_ID': 'SAN1234567',
                        'TAXON_ID': 6344,
                        'SCIENTIFIC_NAME': 'Arenicola marina',
                        'GENUS': 'Arenicola',
                        'FAMILY': 'Arenicolidae',
                        'ORDER_OR_GROUP': 'Scolecida',
                        'COMMON_NAME': 'lugworm',
                        'LIFESTAGE': 'ADULT',
                        'SEX': 'FEMALE',
                        'ORGANISM_PART': 'MUSCLE',
                        'GAL': 'SANGER INSTITUTE',
                        'GAL_SAMPLE_ID': 'SAN0000100',
                        'COLLECTED_BY': 'ALEX COLLECTOR',
                        'COLLECTOR_AFFILIATION': 'THE COLLECTOR INSTITUTE',
                        'DATE_OF_COLLECTION': '2020-09-01',
                        'COLLECTION_LOCATION': 'UNITED KINGDOM | DARK FOREST',
                        'DECIMAL_LATITUDE': '+50.12345678',
                        'DECIMAL_LONGITUDE': '-1.98765432',
                        'HABITAT': 'Woodland',
                        'IDENTIFIED_BY': 'JO IDENTIFIER',
                        'IDENTIFIER_AFFILIATION': 'THE IDENTIFIER INSTITUTE',
                        'VOUCHER_ID': 'voucher1',
                        'OTHER_INFORMATION': None,
                        'DEPTH': None,
                        'ELEVATION': None,
                        'RELATIONSHIP': None,
                        'SYMBIONT': None,
                        'CULTURE_OR_STRAIN_ID': None,
                        'SERIES': None,
                        'RACK_OR_PLATE_ID': None,
                        'TUBE_OR_WELL_ID': None,
                        'TAXON_REMARKS': None,
                        'INFRASPECIFIC_EPITHET': None,
                        'COLLECTOR_SAMPLE_ID': None,
                        'GRID_REFERENCE': None,
                        'TIME_OF_COLLECTION': None,
                        'DESCRIPTION_OF_COLLECTION_METHOD': None,
                        'DIFFICULT_OR_HIGH_PRIORITY_SAMPLE': None,
                        'IDENTIFIED_HOW': None,
                        'SPECIMEN_ID_RISK': None,
                        'PRESERVED_BY': None,
                        'PRESERVER_AFFILIATION': None,
                        'PRESERVATION_APPROACH': None,
                        'PRESERVATIVE_SOLUTION': None,
                        'TIME_ELAPSED_FROM_COLLECTION_TO_PRESERVATION': None,
                        'DATE_OF_PRESERVATION': None,
                        'SIZE_OF_TISSUE_IN_TUBE': None,
                        'TISSUE_REMOVED_FOR_BARCODING': None,
                        'PLATE_ID_FOR_BARCODING': None,
                        'TUBE_OR_WELL_ID_FOR_BARCODING': None,
                        'TISSUE_FOR_BARCODING': None,
                        'BARCODE_PLATE_PRESERVATIVE': None,
                        'PURPOSE_OF_SPECIMEN': None,
                        'HAZARD_GROUP': None,
                        'REGULATORY_COMPLIANCE': None,
                        'ORIGINAL_COLLECTION_DATE': None,
                        'ORIGINAL_GEOGRAPHIC_LOCATION': None,
                        'BARCODE_HUB': None,
                        'tolId': 'wuAreMari1',
                        'biosampleAccession': "SAMEA8521239",
                        'sraAccession': "ERS6206028",
                        'submissionAccession': "ERA3819349",
                        'submissionError': None,
                        'sampleSameAs': None,
                        'sampleDerivedFrom': 'SAMEA12345678',
                        'sampleSymbiontOf': None}
                    ]}
        self.assertEqual(expected, response.json)

    def test_upload_manifest_excel(self):
        # No authorisation token given
        body = []
        response = self.client.open(
            '/api/v1/manifests/upload-excel',
            method='POST',
            json=body)
        self.assert401(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Invalid authorisation token given
        body = []
        response = self.client.open(
            '/api/v1/manifests/upload-excel',
            method='POST',
            headers={"api-key": "12345678"},
            json=body)
        self.assert401(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Excel file missing
        data = {}
        response = self.client.open(
            '/api/v1/manifests/upload-excel',
            method='POST',
            headers={"api-key": self.user3.api_key},
            data=data)
        self.assert400(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Excel file with no taxon ID, specimen ID
        file = open('test/test-manifest-field-missing.xlsx', 'rb')
        data = {
            'excelFile': (file, 'test_file.xlsx'),
        }
        expected = {'manifestId': None,
                    'number_of_errors': 1,
                    'validations': [{'results': [], 'row': 1},
                                    {'results': [{'field': 'SCIENTIFIC_NAME',
                                                  'message': 'A value must be given',
                                                  'severity': 'ERROR'}],
                                     'row': 2}]}
        response = self.client.open(
            '/api/v1/manifests/upload-excel',
            method='POST',
            headers={"api-key": self.user3.api_key},
            data=data)
        file.close()
        self.assert400(response,
                       'Response body is : ' + response.data.decode('utf-8'))
        self.assertEqual(expected, response.json)

        # User not a submitter
        file = open('test/test-manifest.xlsx', 'rb')
        data = {
            'excelFile': (file, 'test_file.xlsx'),
        }
        response = self.client.open(
            '/api/v1/manifests/upload-excel',
            method='POST',
            headers={"api-key": self.user1.api_key},
            data=data)
        file.close()
        self.assert403(response, 'Not received a 403 response')

        # Excel file correct
        file = open('test/test-manifest.xlsx', 'rb')
        data = {
            'excelFile': (file, 'test_file.xlsx'),
        }
        response = self.client.open(
            '/api/v1/manifests/upload-excel',
            method='POST',
            headers={"api-key": self.user3.api_key},
            data=data)
        file.close()
        self.assert200(response, 'Response body is : ' + response.data.decode('utf-8'))
        expected = {'manifestId': 1,
                    'projectName': 'ToL',
                    'stsManifestId': None,
                    'samples': [{
                        'COLLECTED_BY': 'FABIAN HERDER',
                        'COLLECTION_LOCATION': 'Germany | Bonn | Zoological Research Museum Alexander Koenig Leibniz Institute for Animal Biodiversity',  # noqa
                        'COLLECTOR_AFFILIATION': 'ZFMK Adenauerallee 160 53113 Bonn Germany',
                        'COMMON_NAME': None,
                        'CULTURE_OR_STRAIN_ID': None,
                        'DATE_OF_COLLECTION': '2020-12-08',
                        'DECIMAL_LATITUDE': '50.7223215315783',
                        'DECIMAL_LONGITUDE': '7.11382096910185',
                        'DEPTH': None,
                        'ELEVATION': None,
                        'FAMILY': 'Telmatherinidae',
                        'GAL': 'SANGER INSTITUTE',
                        'GAL_SAMPLE_ID': 'SAN2300000',
                        'GENUS': 'Telmatherina',
                        'HABITAT': 'Aquarium ZFMK Bonn',
                        'IDENTIFIED_BY': 'FABIAN HERDER',
                        'IDENTIFIER_AFFILIATION': 'ZFMK Adenauerallee 160 53113 Bonn Germany',
                        'LIFESTAGE': 'ADULT',
                        'ORDER_OR_GROUP': 'Atheriniformes',
                        'ORGANISM_PART': 'TESTIS',
                        'OTHER_INFORMATION': None,
                        'RELATIONSHIP': None,
                        'SCIENTIFIC_NAME': 'Telmatherina bonti',
                        'SEX': 'MALE',
                        'SPECIMEN_ID': 'SAN2300000',
                        'SYMBIONT': None,
                        'TAXON_ID': 446457,
                        'VOUCHER_ID': 'NOT_PROVIDED',
                        'SERIES': None,
                        'RACK_OR_PLATE_ID': None,
                        'TUBE_OR_WELL_ID': None,
                        'TAXON_REMARKS': None,
                        'INFRASPECIFIC_EPITHET': None,
                        'COLLECTOR_SAMPLE_ID': None,
                        'GRID_REFERENCE': None,
                        'TIME_OF_COLLECTION': None,
                        'DESCRIPTION_OF_COLLECTION_METHOD': None,
                        'DIFFICULT_OR_HIGH_PRIORITY_SAMPLE': None,
                        'IDENTIFIED_HOW': None,
                        'SPECIMEN_ID_RISK': None,
                        'PRESERVED_BY': None,
                        'PRESERVER_AFFILIATION': None,
                        'PRESERVATION_APPROACH': None,
                        'PRESERVATIVE_SOLUTION': None,
                        'TIME_ELAPSED_FROM_COLLECTION_TO_PRESERVATION': None,
                        'DATE_OF_PRESERVATION': None,
                        'SIZE_OF_TISSUE_IN_TUBE': None,
                        'TISSUE_REMOVED_FOR_BARCODING': None,
                        'PLATE_ID_FOR_BARCODING': None,
                        'TUBE_OR_WELL_ID_FOR_BARCODING': None,
                        'TISSUE_FOR_BARCODING': None,
                        'BARCODE_PLATE_PRESERVATIVE': None,
                        'PURPOSE_OF_SPECIMEN': None,
                        'HAZARD_GROUP': None,
                        'REGULATORY_COMPLIANCE': None,
                        'ORIGINAL_COLLECTION_DATE': None,
                        'ORIGINAL_GEOGRAPHIC_LOCATION': None,
                        'BARCODE_HUB': None,
                        'biosampleAccession': None,
                        'row': 1,
                        'sampleDerivedFrom': None,
                        'sampleSameAs': None,
                        'sampleSymbiontOf': None,
                        'sraAccession': None,
                        'submissionAccession': None,
                        'submissionError': None,
                        'tolId': None}, {
                        'COLLECTED_BY': 'FABIAN HERDER',
                        'COLLECTION_LOCATION': 'Germany | Bonn | Zoological Research Museum Alexander Koenig Leibniz Institute for Animal Biodiversity',  # noqa
                        'COLLECTOR_AFFILIATION': 'ZFMK Adenauerallee 160 53113 Bonn Germany',
                        'COMMON_NAME': None,
                        'CULTURE_OR_STRAIN_ID': None,
                        'DATE_OF_COLLECTION': '2020-12-08',
                        'DECIMAL_LATITUDE': '50.7223215315783',
                        'DECIMAL_LONGITUDE': '7.11382096910185',
                        'DEPTH': None,
                        'ELEVATION': None,
                        'FAMILY': 'Telmatherinidae',
                        'GAL': 'SANGER INSTITUTE',
                        'GAL_SAMPLE_ID': 'SAN2300000',
                        'GENUS': 'Telmatherina',
                        'HABITAT': 'Aquarium ZFMK Bonn',
                        'IDENTIFIED_BY': 'FABIAN HERDER',
                        'IDENTIFIER_AFFILIATION': 'ZFMK Adenauerallee 160 53113 Bonn Germany',
                        'LIFESTAGE': 'ADULT',
                        'ORDER_OR_GROUP': 'Atheriniformes',
                        'ORGANISM_PART': 'LIVER',
                        'OTHER_INFORMATION': None,
                        'RELATIONSHIP': None,
                        'SCIENTIFIC_NAME': 'Telmatherina bonti',
                        'SEX': 'MALE',
                        'SPECIMEN_ID': 'SAN2300000',
                        'SYMBIONT': None,
                        'TAXON_ID': 446457,
                        'VOUCHER_ID': 'NOT_PROVIDED',
                        'SERIES': None,
                        'RACK_OR_PLATE_ID': None,
                        'TUBE_OR_WELL_ID': None,
                        'TAXON_REMARKS': None,
                        'INFRASPECIFIC_EPITHET': None,
                        'COLLECTOR_SAMPLE_ID': None,
                        'GRID_REFERENCE': None,
                        'TIME_OF_COLLECTION': None,
                        'DESCRIPTION_OF_COLLECTION_METHOD': None,
                        'DIFFICULT_OR_HIGH_PRIORITY_SAMPLE': None,
                        'IDENTIFIED_HOW': None,
                        'SPECIMEN_ID_RISK': None,
                        'PRESERVED_BY': None,
                        'PRESERVER_AFFILIATION': None,
                        'PRESERVATION_APPROACH': None,
                        'PRESERVATIVE_SOLUTION': None,
                        'TIME_ELAPSED_FROM_COLLECTION_TO_PRESERVATION': None,
                        'DATE_OF_PRESERVATION': None,
                        'SIZE_OF_TISSUE_IN_TUBE': None,
                        'TISSUE_REMOVED_FOR_BARCODING': None,
                        'PLATE_ID_FOR_BARCODING': None,
                        'TUBE_OR_WELL_ID_FOR_BARCODING': None,
                        'TISSUE_FOR_BARCODING': None,
                        'BARCODE_PLATE_PRESERVATIVE': None,
                        'PURPOSE_OF_SPECIMEN': None,
                        'HAZARD_GROUP': None,
                        'REGULATORY_COMPLIANCE': None,
                        'ORIGINAL_COLLECTION_DATE': None,
                        'ORIGINAL_GEOGRAPHIC_LOCATION': None,
                        'BARCODE_HUB': None,
                        'biosampleAccession': None,
                        'row': 2,
                        'sampleDerivedFrom': None,
                        'sampleSameAs': None,
                        'sampleSymbiontOf': None,
                        'sraAccession': None,
                        'submissionAccession': None,
                        'submissionError': None,
                        'tolId': None}],
                    'submissionStatus': None}
        self.assertEqual(expected, response.json)
        # Has the Excel file been uploaded?
        manifest = db.session.query(SubmissionsManifest) \
            .filter(SubmissionsManifest.manifest_id == 1) \
            .one_or_none()
        self.assertEqual("1.xlsx", manifest.excel_file)

        # Get that from minio and check it is the same file as went in
        dir = tempfile.TemporaryDirectory()
        excel_utils.load_excel(manifest=manifest, dirname=dir.name, filename="test.xlsx")
        self.assertTrue(filecmp.cmp(dir.name + "/test.xlsx",
                                    'test/test-manifest.xlsx',
                                    shallow=False))

    def test_download_manifest_excel(self):
        # Upload it first
        file = open('test/test-manifest.xlsx', 'rb')
        data = {
            'excelFile': (file, 'test_file.xlsx'),
        }
        response = self.client.open(
            '/api/v1/manifests/upload-excel',
            method='POST',
            headers={"api-key": self.user3.api_key},
            data=data)
        file.close()
        self.assert200(response, 'Response body is : ' + response.data.decode('utf-8'))
        # Don't need to test the response here as it has been tested before

        # Pretend all the ID columns have been generated
        manifest = db.session.query(SubmissionsManifest) \
            .filter(SubmissionsManifest.manifest_id == 1) \
            .one_or_none()
        counter = 1
        for sample in manifest.samples:
            for field in SubmissionsSample.id_fields:
                setattr(sample, field["python_name"], str(counter))
                counter += 1
        # Make sure there is a None in there
        manifest.samples[0].sample_symbiont_of = None

        # No authorisation token given
        response = self.client.open(
            '/api/v1/manifests/1/download-excel',
            method='GET')
        self.assert401(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Invalid authorisation token given
        response = self.client.open(
            '/api/v1/manifests/1/download-excel',
            method='GET',
            headers={"api-key": "12345678"})
        self.assert401(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Manifest missing
        response = self.client.open(
            '/api/v1/manifests/999/download-excel',
            method='GET',
            headers={"api-key": self.user3.api_key})
        self.assert404(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # User not a submitter
        response = self.client.open(
            '/api/v1/manifests/1/download-excel',
            method='GET',
            headers={"api-key": self.user1.api_key})
        file.close()
        self.assert403(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        response = self.client.open(
            '/api/v1/manifests/1/download-excel',
            method='GET',
            headers={"api-key": self.user3.api_key})
        file.close()
        self.assert200(response,
                       'Did not receive a 200 response')
        self.assertEqual('application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         response.content_type)

        # Save as Excel file
        file = open('test/test-manifest-validated.xlsx', 'wb')
        file.write(response.get_data())
        file.close()
        workbook = load_workbook(filename='test/test-manifest-validated.xlsx')
        sheet = workbook.active
        # Do we have the columns filled in that we expect
        for sample in manifest.samples:
            for field in SubmissionsSample.id_fields:
                column = excel_utils.find_column(sheet, field["field_name"]) + 1
                row = sample.row + 1
                value = getattr(sample, field["python_name"])
                self.assertEqual(sheet.cell(row=row, column=column).value, value)

    def test_download_manifest_excel_originally_json(self):
        body = {'samples': [
                    {'row': 1,
                     'SPECIMEN_ID': 'SAN1234567',
                     'TAXON_ID': 6344,
                     'SCIENTIFIC_NAME': 'Arenicola marina',
                     'GENUS': 'Arenicola',
                     'FAMILY': 'Arenicolidae',
                     'ORDER_OR_GROUP': 'Scolecida',
                     'COMMON_NAME': 'lugworm',
                     'LIFESTAGE': 'ADULT',
                     'SEX': 'FEMALE',
                     'ORGANISM_PART': 'MUSCLE',
                     'GAL': 'SANGER INSTITUTE',
                     'GAL_SAMPLE_ID': 'SAN0000100',
                     'COLLECTED_BY': 'ALEX COLLECTOR',
                     'COLLECTOR_AFFILIATION': 'THE COLLECTOR INSTITUTE',
                     'DATE_OF_COLLECTION': '2020-09-01',
                     'COLLECTION_LOCATION': 'UNITED KINGDOM | DARK FOREST',
                     'DECIMAL_LATITUDE': '+50.12345678',
                     'DECIMAL_LONGITUDE': '-1.98765432',
                     'HABITAT': 'Woodland',
                     'IDENTIFIED_BY': 'JO IDENTIFIER',
                     'IDENTIFIER_AFFILIATION': 'THE IDENTIFIER INSTITUTE',
                     'VOUCHER_ID': 'voucher1'}
                ]}

        # Correct, full JSON
        response = self.client.open(
            '/api/v1/manifests',
            method='POST',
            headers={"api-key": self.user3.api_key},
            json=body)
        self.assert200(response,
                       'Response body is : ' + response.data.decode('utf-8'))
        # Don't need to test the response here as it has been tested before

        # Pretend all the ID columns have been generated
        manifest = db.session.query(SubmissionsManifest) \
            .filter(SubmissionsManifest.manifest_id == 1) \
            .one_or_none()
        counter = 1
        for sample in manifest.samples:
            for field in SubmissionsSample.id_fields:
                setattr(sample, field["python_name"], str(counter))
                counter += 1
        # Make sure there is a None in there
        manifest.samples[0].sample_symbiont_of = None

        response = self.client.open(
            '/api/v1/manifests/1/download-excel',
            method='GET',
            headers={"api-key": self.user3.api_key})
        self.assert200(response,
                       'Did not receive a 200 response')
        self.assertEqual('application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         response.content_type)

        # Save as Excel file
        file = open('test/test-manifest-validated.xlsx', 'wb')
        file.write(response.get_data())
        file.close()
        workbook = load_workbook(filename='test/test-manifest-validated.xlsx')
        sheet = workbook.active
        # Do we have the columns filled in that we expect
        for sample in manifest.samples:
            for field in SubmissionsSample.id_fields:
                column = excel_utils.find_column(sheet, field["field_name"]) + 1
                row = sample.row + 1
                value = getattr(sample, field["python_name"])
                self.assertEqual(sheet.cell(row=row, column=column).value, value)

        # Do we have the original columns in?
        for sample in manifest.samples:
            for field in SubmissionsSample.all_fields:
                column = excel_utils.find_column(sheet, field["field_name"]) + 1
                row = sample.row + 1
                value = getattr(sample, field["python_name"])
                self.assertEqual(sheet.cell(row=row, column=column).value, value)


if __name__ == '__main__':
    import unittest
    unittest.main()
