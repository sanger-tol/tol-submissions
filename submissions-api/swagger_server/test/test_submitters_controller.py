from __future__ import absolute_import

from swagger_server.test import BaseTestCase
from swagger_server.model import db, SubmissionsManifest, SubmissionsSpecimen, \
    SubmissionsSample
import swagger_server.excel_utils as excel_utils
import os
import responses
import tempfile
import filecmp
import datetime
from openpyxl import load_workbook


class TestSubmittersController(BaseTestCase):

    def test_submit_manifest_json(self):
        # No authorisation token given
        body = []
        response = self.client.open(
            '/api/v1/manifests',
            method='POST',
            json=body)
        self.assert401(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Invalid authorisation token given
        body = []
        response = self.client.open(
            '/api/v1/manifests',
            method='POST',
            headers={"api-key": "12345678"},
            json=body)
        self.assert401(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Body empty
        body = {}
        response = self.client.open(
            '/api/v1/manifests',
            method='POST',
            headers={"api-key": self.user3.api_key},
            json=body)
        self.assert400(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        body = {'samples': [
                    {'row': 1,
                     'SPECIMEN_ID': 'specimen9876',
                     'TAXON_ID': 6344,
                     'SCIENTIFIC_NAME': 'Arenicola marina',
                     'GENUS': 'Arenicola',
                     'FAMILY': 'Arenicolidae',
                     'ORDER_OR_GROUP': 'None',
                     'COMMON_NAME': 'lugworm',
                     'LIFESTAGE': 'ADULT',
                     'SEX': 'FEMALE',
                     'ORGANISM_PART': 'MUSCLE',
                     'GAL': 'SANGER INSTITUTE',
                     'GAL_SAMPLE_ID': 'SAN000100',
                     'COLLECTED_BY': 'ALEX COLLECTOR',
                     'COLLECTOR_AFFILIATION': 'THE COLLECTOR INSTITUTE',
                     'DATE_OF_COLLECTION': '2020-09-01',
                     'COLLECTION_LOCATION': 'UNITED KINGDOM | DARK FOREST',
                     'DECIMAL_LATITUDE': '+50.12345678',
                     'DECIMAL_LONGITUDE': '-1.98765432',
                     'HABITAT': 'Woodland',
                     'IDENTIFIED_BY': 'JO IDENTIFIER',
                     'IDENTIFIER_AFFILIATION': 'THE IDENTIFIER INSTITUTE',
                     'VOUCHER_ID': 'voucher1'}
                ]}
        # Not a submitter
        response = self.client.open(
            '/api/v1/manifests',
            method='POST',
            headers={"api-key": self.user1.api_key},
            json=body)
        self.assert403(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Correct, full JSON
        response = self.client.open(
            '/api/v1/manifests',
            method='POST',
            headers={"api-key": self.user3.api_key},
            json=body)
        self.assert200(response,
                       'Response body is : ' + response.data.decode('utf-8'))
        expected = {'manifestId': 1,
                    'submissionStatus': None,
                    'projectName': 'ToL',  # Default - not given in body
                    'stsManifestId': None,  # Default - not given in body
                    'samples': [
                        {'row': 1,
                         'SPECIMEN_ID': 'specimen9876',
                         'TAXON_ID': 6344,
                         'SCIENTIFIC_NAME': 'Arenicola marina',
                         'GENUS': 'Arenicola',
                         'FAMILY': 'Arenicolidae',
                         'ORDER_OR_GROUP': 'None',
                         'COMMON_NAME': 'lugworm',
                         'LIFESTAGE': 'ADULT',
                         'SEX': 'FEMALE',
                         'ORGANISM_PART': 'MUSCLE',
                         'GAL': 'SANGER INSTITUTE',
                         'GAL_SAMPLE_ID': 'SAN000100',
                         'COLLECTED_BY': 'ALEX COLLECTOR',
                         'COLLECTOR_AFFILIATION': 'THE COLLECTOR INSTITUTE',
                         'DATE_OF_COLLECTION': '2020-09-01',
                         'COLLECTION_LOCATION': 'UNITED KINGDOM | DARK FOREST',
                         'DECIMAL_LATITUDE': '+50.12345678',
                         'DECIMAL_LONGITUDE': '-1.98765432',
                         'HABITAT': 'Woodland',
                         'IDENTIFIED_BY': 'JO IDENTIFIER',
                         'IDENTIFIER_AFFILIATION': 'THE IDENTIFIER INSTITUTE',
                         'VOUCHER_ID': 'voucher1',
                         'ELEVATION': None,
                         'DEPTH': None,
                         'RELATIONSHIP': None,
                         'SYMBIONT': None,
                         'CULTURE_OR_STRAIN_ID': None,
                         'tolId': None,
                         'biosampleAccession': None,
                         'sraAccession': None,
                         'submissionAccession': None,
                         'submissionError': None,
                         'sampleSameAs': None,
                         'sampleDerivedFrom': None,
                         'sampleSymbiontOf': None}
                    ]}
        self.assertEqual(expected, response.json)

    def test_get_manifest(self):

        body = {'samples': [
                    {'row': 1,
                     'SPECIMEN_ID': 'specimen9876',
                     'TAXON_ID': 6344,
                     'SCIENTIFIC_NAME': 'Arenicola marina',
                     'GENUS': 'Arenicola',
                     'FAMILY': 'Arenicolidae',
                     'ORDER_OR_GROUP': 'None',
                     'COMMON_NAME': 'lugworm',
                     'LIFESTAGE': 'ADULT',
                     'SEX': 'FEMALE',
                     'ORGANISM_PART': 'MUSCLE',
                     'GAL': 'SANGER INSTITUTE',
                     'GAL_SAMPLE_ID': 'SAN000100',
                     'COLLECTED_BY': 'ALEX COLLECTOR',
                     'COLLECTOR_AFFILIATION': 'THE COLLECTOR INSTITUTE',
                     'DATE_OF_COLLECTION': '2020-09-01',
                     'COLLECTION_LOCATION': 'UNITED KINGDOM | DARK FOREST',
                     'DECIMAL_LATITUDE': '+50.12345678',
                     'DECIMAL_LONGITUDE': '-1.98765432',
                     'HABITAT': 'Woodland',
                     'IDENTIFIED_BY': 'JO IDENTIFIER',
                     'IDENTIFIER_AFFILIATION': 'THE IDENTIFIER INSTITUTE',
                     'VOUCHER_ID': 'voucher1'}
                ],
                'projectName': 'TestProj',
                'stsManifestId': '1234-4321'}

        # Submit the manifest
        response = self.client.open(
            '/api/v1/manifests',
            method='POST',
            headers={"api-key": self.user3.api_key},
            json=body)
        self.assert200(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # No authorisation token given
        body = []
        response = self.client.open(
            '/api/v1/manifests/1',
            method='GET',
            json=body)
        self.assert401(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Invalid authorisation token given
        body = []
        response = self.client.open(
            '/api/v1/manifests/1',
            method='GET',
            headers={"api-key": "12345678"},
            json=body)
        self.assert401(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Incorrect manifest ID
        body = {}
        response = self.client.open(
            '/api/v1/manifests/2',
            method='GET',
            headers={"api-key": self.user3.api_key},
            json=body)
        self.assert404(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Not a submitter
        response = self.client.open(
            '/api/v1/manifests/1',
            method='GET',
            headers={"api-key": self.user1.api_key},
            json=body)
        self.assert403(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Corect - should validate and return errors
        response = self.client.open(
            '/api/v1/manifests/1',
            method='GET',
            headers={"api-key": self.user3.api_key},
            json=body)
        self.assert200(response,
                       'Response body is : ' + response.data.decode('utf-8'))
        expected = {'manifestId': 1,
                    'submissionStatus': None,
                    'projectName': 'TestProj',
                    'stsManifestId': '1234-4321',
                    'samples': [{
                        'row': 1,
                        'SPECIMEN_ID': 'specimen9876',
                        'TAXON_ID': 6344,
                        'SCIENTIFIC_NAME': 'Arenicola marina',
                        'GENUS': 'Arenicola',
                        'FAMILY': 'Arenicolidae',
                        'ORDER_OR_GROUP': 'None',
                        'COMMON_NAME': 'lugworm',
                        'LIFESTAGE': 'ADULT',
                        'SEX': 'FEMALE',
                        'ORGANISM_PART': 'MUSCLE',
                        'GAL': 'SANGER INSTITUTE',
                        'GAL_SAMPLE_ID': 'SAN000100',
                        'COLLECTED_BY': 'ALEX COLLECTOR',
                        'COLLECTOR_AFFILIATION': 'THE COLLECTOR INSTITUTE',
                        'DATE_OF_COLLECTION': '2020-09-01',
                        'COLLECTION_LOCATION': 'UNITED KINGDOM | DARK FOREST',
                        'DECIMAL_LATITUDE': '+50.12345678',
                        'DECIMAL_LONGITUDE': '-1.98765432',
                        'HABITAT': 'Woodland',
                        'IDENTIFIED_BY': 'JO IDENTIFIER',
                        'IDENTIFIER_AFFILIATION': 'THE IDENTIFIER INSTITUTE',
                        'VOUCHER_ID': 'voucher1',
                        'DEPTH': None,
                        'ELEVATION': None,
                        'RELATIONSHIP': None,
                        'SYMBIONT': None,
                        'CULTURE_OR_STRAIN_ID': None,
                        'tolId': None,
                        'biosampleAccession': None,
                        'sraAccession': None,
                        'submissionAccession': None,
                        'submissionError': None,
                        'sampleSameAs': None,
                        'sampleDerivedFrom': None,
                        'sampleSymbiontOf': None}

                    ]}
        self.assertEqual(expected, response.json)

    def test_get_manifests(self):

        manifest1 = SubmissionsManifest()
        manifest1.sts_manifest_id = "123-456-789"
        manifest1.project_name = "WowProj"
        manifest1.submission_status = True
        manifest1.created_at = datetime.datetime(2020, 5, 17)
        manifest1.user = self.user1
        db.session.add(manifest1)
        manifest2 = SubmissionsManifest()
        manifest2.sts_manifest_id = "987-654-321"
        manifest2.project_name = "DudProj"
        manifest2.submission_status = False
        manifest2.created_at = datetime.datetime(2021, 6, 18)
        manifest2.user = self.user2
        sample1 = SubmissionsSample(collected_by="ALEX COLLECTOR",
                                    collection_location="UNITED KINGDOM | DARK FOREST",
                                    collector_affiliation="THE COLLECTOR INSTITUTE",
                                    common_name="lugworm",
                                    date_of_collection="2020-09-01",
                                    decimal_latitude="50.12345678",
                                    decimal_longitude="-1.98765432",
                                    depth="100",
                                    elevation="0",
                                    family="Arenicolidae",
                                    GAL="SANGER INSTITUTE",
                                    GAL_sample_id="SAN000100",
                                    genus="Arenicola",
                                    habitat="Woodland",
                                    identified_by="JO IDENTIFIER",
                                    identifier_affiliation="THE IDENTIFIER INSTITUTE",
                                    lifestage="ADULT",
                                    organism_part="MUSCLE",
                                    order_or_group="None",
                                    relationship="child of SAMEA1234567",
                                    scientific_name="Arenicola marina",
                                    sex="FEMALE",
                                    specimen_id="SAN000100",
                                    taxonomy_id=6344,
                                    voucher_id="voucher1",
                                    row=1)
        sample1.manifest = manifest1
        db.session.add(manifest2)
        db.session.commit()

        # No authorisation token given
        body = []
        response = self.client.open(
            '/api/v1/manifests',
            method='GET',
            json=body)
        self.assert401(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Invalid authorisation token given
        body = []
        response = self.client.open(
            '/api/v1/manifests',
            method='GET',
            headers={"api-key": "12345678"},
            json=body)
        self.assert401(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Not a submitter
        response = self.client.open(
            '/api/v1/manifests',
            method='GET',
            headers={"api-key": self.user1.api_key},
            json=body)
        self.assert403(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Correct - should validate and return errors
        response = self.client.open(
            '/api/v1/manifests',
            method='GET',
            headers={"api-key": self.user3.api_key},
            json=body)
        self.assert200(response,
                       'Response body is : ' + response.data.decode('utf-8'))
        expected = [{'manifestId': 2,
                     'submissionStatus': False,
                     'projectName': 'DudProj',
                     'stsManifestId': '987-654-321',
                     'createdAt': '2021-06-18T00:00:00Z',
                     'numberOfSamples': 0,
                     'user': {'email': 'test_user_admin@sanger.ac.uk',
                              'name': 'test_user_admin',
                              'organisation': 'Sanger Institute',
                              'roles': [{'role': 'admin'}]}},
                    {'manifestId': 1,
                     'submissionStatus': True,
                     'projectName': 'WowProj',
                     'stsManifestId': '123-456-789',
                     'createdAt': '2020-05-17T00:00:00Z',
                     'numberOfSamples': 1,
                     'user': {'email': 'test_user_requester@sanger.ac.uk',
                              'name': 'test_user_requester',
                              'organisation': 'Sanger Institute',
                              'roles': []}}
                    ]
        self.assertEqual(expected, response.json)

    @responses.activate
    def test_validate_manifest_json(self):
        mock_response_from_ena = {"taxId": "6344",
                                  "scientificName": "Arenicola marina",
                                  "commonName": "lugworm",
                                  "formalName": "true",
                                  "rank": "species",
                                  "division": "INV",
                                  "lineage": "",
                                  "geneticCode": "1",
                                  "mitochondrialGeneticCode": "5",
                                  "submittable": "true"}
        responses.add(responses.GET, 'https://www.ebi.ac.uk/ena/taxonomy/rest/tax-id/6344',
                      json=mock_response_from_ena, status=200)
        mock_response_from_tolid = [{"taxonomyId": "6344",
                                     "scientificName": "Arenicola marina",
                                     "commonName": "lugworm",
                                     "family": "Arenicolidae",
                                     "genus": "Arenicola",
                                     "order": "None"}]
        responses.add(responses.GET, os.environ['TOLID_URL'] + '/species/6344',
                      json=mock_response_from_tolid, status=200)

        body = {'samples': [
                    {'row': 1,
                     'SPECIMEN_ID': 'specimen9876',
                     'TAXON_ID': 6344,
                     'SCIENTIFIC_NAME': 'Arenicola marina',
                     'GENUS': 'Arenicola',
                     'FAMILY': 'Arenicolidae',
                     'ORDER_OR_GROUP': 'None',
                     'COMMON_NAME': 'lugworm',
                     'LIFESTAGE': 'ADULT',
                     'SEX': 'FEMALE',
                     'ORGANISM_PART': '',
                     'GAL': 'SANGER INSTITUTE',
                     'GAL_SAMPLE_ID': 'SAN000100',
                     'COLLECTED_BY': 'ALEX COLLECTOR',
                     'COLLECTOR_AFFILIATION': 'THE COLLECTOR INSTITUTE',
                     'DATE_OF_COLLECTION': '2020-09-01',
                     'COLLECTION_LOCATION': 'UNITED KINGDOM | DARK FOREST',
                     'DECIMAL_LATITUDE': '+50.12345678',
                     'DECIMAL_LONGITUDE': '-1.98765432',
                     'HABITAT': 'Woodland',
                     'IDENTIFIED_BY': 'JO IDENTIFIER',
                     'IDENTIFIER_AFFILIATION': 'THE IDENTIFIER INSTITUTE',
                     'VOUCHER_ID': 'voucher1'}
                ]}

        # Submit the manifest
        response = self.client.open(
            '/api/v1/manifests',
            method='POST',
            headers={"api-key": self.user3.api_key},
            json=body)
        self.assert200(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # No authorisation token given
        body = []
        response = self.client.open(
            '/api/v1/manifests/1/validate',
            method='GET',
            json=body)
        self.assert401(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Invalid authorisation token given
        body = []
        response = self.client.open(
            '/api/v1/manifests/1/validate',
            method='GET',
            headers={"api-key": "12345678"},
            json=body)
        self.assert401(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Incorrect manifest ID
        body = {}
        response = self.client.open(
            '/api/v1/manifests/2/validate',
            method='GET',
            headers={"api-key": self.user3.api_key},
            json=body)
        self.assert404(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Not a submitter
        response = self.client.open(
            '/api/v1/manifests/1/validate',
            method='GET',
            headers={"api-key": self.user1.api_key},
            json=body)
        self.assert403(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Corect - should validate and return errors
        response = self.client.open(
            '/api/v1/manifests/1/validate',
            method='GET',
            headers={"api-key": self.user3.api_key},
            json=body)
        self.assert200(response,
                       'Response body is : ' + response.data.decode('utf-8'))
        expected = {'manifestId': 1,
                    "number_of_errors": 1,
                    'validations': [
                        {"row": 1,
                         "results": [
                             {"field": "ORGANISM_PART",
                              "message": "Must not be empty",
                              'severity': 'ERROR'}
                         ]}
                    ]}
        self.assertEqual(expected, response.json)

    @responses.activate
    def test_submit_and_validate_manifest_json(self):
        mock_response_from_ena = {"taxId": "6344",
                                  "scientificName": "Arenicola marina",
                                  "commonName": "lugworm",
                                  "formalName": "true",
                                  "rank": "species",
                                  "division": "INV",
                                  "lineage": "",
                                  "geneticCode": "1",
                                  "mitochondrialGeneticCode": "5",
                                  "submittable": "true"}
        responses.add(responses.GET, 'https://www.ebi.ac.uk/ena/taxonomy/rest/tax-id/6344',
                      json=mock_response_from_ena, status=200)
        mock_response_from_tolid = [{"taxonomyId": "6344",
                                     "scientificName": "Arenicola marina",
                                     "commonName": "lugworm",
                                     "family": "Arenicolidae",
                                     "genus": "Arenicola",
                                     "order": "None"}]
        responses.add(responses.GET, os.environ['TOLID_URL'] + '/species/6344',
                      json=mock_response_from_tolid, status=200)

        # No authorisation token given
        body = []
        response = self.client.open(
            '/api/v1/manifests/validate',
            method='POST',
            json=body)
        self.assert401(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Invalid authorisation token given
        body = []
        response = self.client.open(
            '/api/v1/manifests/validate',
            method='POST',
            headers={"api-key": "12345678"},
            json=body)
        self.assert401(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Body empty
        body = {}
        response = self.client.open(
            '/api/v1/manifests/validate',
            method='POST',
            headers={"api-key": self.user3.api_key},
            json=body)
        self.assert400(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        body = {'samples': [
                    {'row': 1,
                     'SPECIMEN_ID': 'specimen9876',
                     'TAXON_ID': 6344,
                     'SCIENTIFIC_NAME': 'Arenicola marina',
                     'GENUS': 'Arenicola',
                     'FAMILY': 'Arenicolidae',
                     'ORDER_OR_GROUP': 'None',
                     'COMMON_NAME': 'lugworm',
                     'LIFESTAGE': 'ADULT',
                     'SEX': 'FEMALE',
                     'ORGANISM_PART': '',
                     'GAL': 'SANGER INSTITUTE',
                     'GAL_SAMPLE_ID': 'SAN000100',
                     'COLLECTED_BY': 'ALEX COLLECTOR',
                     'COLLECTOR_AFFILIATION': 'THE COLLECTOR INSTITUTE',
                     'DATE_OF_COLLECTION': '2020-09-01',
                     'COLLECTION_LOCATION': 'UNITED KINGDOM | DARK FOREST',
                     'DECIMAL_LATITUDE': '+50.12345678',
                     'DECIMAL_LONGITUDE': '-1.98765432',
                     'HABITAT': 'Woodland',
                     'IDENTIFIED_BY': 'JO IDENTIFIER',
                     'IDENTIFIER_AFFILIATION': 'THE IDENTIFIER INSTITUTE',
                     'VOUCHER_ID': 'voucher1'}
                ]}
        # Not a submitter
        response = self.client.open(
            '/api/v1/manifests/validate',
            method='POST',
            headers={"api-key": self.user1.api_key},
            json=body)
        self.assert403(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Correct, full JSON
        response = self.client.open(
            '/api/v1/manifests/validate',
            method='POST',
            headers={"api-key": self.user3.api_key},
            json=body)
        self.assert200(response,
                       'Response body is : ' + response.data.decode('utf-8'))
        expected = {'manifestId': 1,
                    "number_of_errors": 1,
                    'validations': [
                        {"row": 1,
                         "results": [
                             {"field": "ORGANISM_PART",
                              "message": "Must not be empty",
                              'severity': 'ERROR'}
                         ]}
                    ]}
        self.assertEqual(expected, response.json)

    @responses.activate
    def test_generate_ids(self):
        specimen = SubmissionsSpecimen()
        specimen.specimen_id = "specimen9876"
        specimen.biosample_accession = "SAMEA12345678"
        db.session.add(specimen)
        db.session.commit()

        mock_response_from_ena = {"taxId": "6344",
                                  "scientificName": "Arenicola marina",
                                  "commonName": "lugworm",
                                  "formalName": "true",
                                  "rank": "species",
                                  "division": "INV",
                                  "lineage": "",
                                  "geneticCode": "1",
                                  "mitochondrialGeneticCode": "5",
                                  "submittable": "true"}
        responses.add(responses.GET, 'https://www.ebi.ac.uk/ena/taxonomy/rest/tax-id/6344',
                      json=mock_response_from_ena, status=200)
        mock_response_from_tolid = [{
            "species": {
                "commonName": "lugworm",
                "family": "Arenicolidae",
                "genus": "Arenicola",
                "kingdom": "Metazoa",
                "order": "Capitellida",
                "phylum": "Annelida",
                "prefix": "wuAreMari",
                "scientificName": "Arenicola marina",
                "taxaClass": "Polychaeta",
                "taxonomyId": 6344
            },
            "specimen": {
                "specimenId": "specimen9876"
            },
            "tolId": "wuAreMari1"
        }]
        responses.add(responses.POST, os.environ['TOLID_URL'] + '/tol-ids',
                      json=mock_response_from_tolid, status=200)

        body = {'samples': [
                    {'row': 1,
                     'SPECIMEN_ID': 'specimen9876',
                     'TAXON_ID': 6344,
                     'SCIENTIFIC_NAME': 'Arenicola marina',
                     'GENUS': 'Arenicola',
                     'FAMILY': 'Arenicolidae',
                     'ORDER_OR_GROUP': 'None',
                     'COMMON_NAME': 'lugworm',
                     'LIFESTAGE': 'ADULT',
                     'SEX': 'FEMALE',
                     'ORGANISM_PART': 'MUSCLE',
                     'GAL': 'SANGER INSTITUTE',
                     'GAL_SAMPLE_ID': 'SAN000100',
                     'COLLECTED_BY': 'ALEX COLLECTOR',
                     'COLLECTOR_AFFILIATION': 'THE COLLECTOR INSTITUTE',
                     'DATE_OF_COLLECTION': '2020-09-01',
                     'COLLECTION_LOCATION': 'UNITED KINGDOM | DARK FOREST',
                     'DECIMAL_LATITUDE': '+50.12345678',
                     'DECIMAL_LONGITUDE': '-1.98765432',
                     'HABITAT': 'Woodland',
                     'IDENTIFIED_BY': 'JO IDENTIFIER',
                     'IDENTIFIER_AFFILIATION': 'THE IDENTIFIER INSTITUTE',
                     'VOUCHER_ID': 'voucher1'}
                ]}

        # Submit the manifest
        response = self.client.open(
            '/api/v1/manifests',
            method='POST',
            headers={"api-key": self.user3.api_key},
            json=body)
        self.assert200(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        manifest = db.session.query(SubmissionsManifest) \
            .filter(SubmissionsManifest.manifest_id == 1) \
            .one_or_none()
        mock_response_from_ena = '<?xml version="1.0" encoding="UTF-8"?><?xml-stylesheet type="text/xsl" href="receipt.xsl"?><RECEIPT receiptDate="2021-04-07T12:47:39.998+01:00" submissionFile="tmphz9luulrsubmission_3.xml" success="true"><SAMPLE accession="ERS6206028" alias="' + str(manifest.samples[0].sample_id) + '" status="PRIVATE"><EXT_ID accession="SAMEA8521239" type="biosample"/></SAMPLE><SUBMISSION accession="ERA3819349" alias="SUBMISSION-07-04-2021-12:47:36:825"/><ACTIONS>ADD</ACTIONS></RECEIPT>'  # noqa
        responses.add(responses.POST, os.environ['ENA_URL'] + '/ena/submit/drop-box/submit/',
                      body=mock_response_from_ena, status=200)

        # No authorisation token given
        body = []
        response = self.client.open(
            '/api/v1/manifests/1/generate',
            method='PATCH',
            json=body)
        self.assert401(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Invalid authorisation token given
        body = []
        response = self.client.open(
            '/api/v1/manifests/1/generate',
            method='PATCH',
            headers={"api-key": "12345678"},
            json=body)
        self.assert401(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Incorrect manifest ID
        body = {}
        response = self.client.open(
            '/api/v1/manifests/2/generate',
            method='PATCH',
            headers={"api-key": self.user3.api_key},
            json=body)
        self.assert404(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Not a submitter
        response = self.client.open(
            '/api/v1/manifests/1/generate',
            method='PATCH',
            headers={"api-key": self.user1.api_key},
            json=body)
        self.assert403(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Corect - should validate and return errors
        response = self.client.open(
            '/api/v1/manifests/1/generate',
            method='PATCH',
            headers={"api-key": self.user3.api_key},
            json=body)
        self.assert200(response,
                       'Response body is : ' + response.data.decode('utf-8'))
        expected = {'manifestId': 1,
                    'submissionStatus': True,
                    'projectName': 'ToL',
                    'stsManifestId': None,
                    'samples': [{
                        'row': 1,
                        'SPECIMEN_ID': 'specimen9876',
                        'TAXON_ID': 6344,
                        'SCIENTIFIC_NAME': 'Arenicola marina',
                        'GENUS': 'Arenicola',
                        'FAMILY': 'Arenicolidae',
                        'ORDER_OR_GROUP': 'None',
                        'COMMON_NAME': 'lugworm',
                        'LIFESTAGE': 'ADULT',
                        'SEX': 'FEMALE',
                        'ORGANISM_PART': 'MUSCLE',
                        'GAL': 'SANGER INSTITUTE',
                        'GAL_SAMPLE_ID': 'SAN000100',
                        'COLLECTED_BY': 'ALEX COLLECTOR',
                        'COLLECTOR_AFFILIATION': 'THE COLLECTOR INSTITUTE',
                        'DATE_OF_COLLECTION': '2020-09-01',
                        'COLLECTION_LOCATION': 'UNITED KINGDOM | DARK FOREST',
                        'DECIMAL_LATITUDE': '+50.12345678',
                        'DECIMAL_LONGITUDE': '-1.98765432',
                        'HABITAT': 'Woodland',
                        'IDENTIFIED_BY': 'JO IDENTIFIER',
                        'IDENTIFIER_AFFILIATION': 'THE IDENTIFIER INSTITUTE',
                        'VOUCHER_ID': 'voucher1',
                        'DEPTH': None,
                        'ELEVATION': None,
                        'RELATIONSHIP': None,
                        'SYMBIONT': None,
                        'CULTURE_OR_STRAIN_ID': None,
                        'tolId': 'wuAreMari1',
                        'biosampleAccession': "SAMEA8521239",
                        'sraAccession': "ERS6206028",
                        'submissionAccession': "ERA3819349",
                        'submissionError': None,
                        'sampleSameAs': None,
                        'sampleDerivedFrom': 'SAMEA12345678',
                        'sampleSymbiontOf': None}
                    ]}
        self.assertEqual(expected, response.json)

    @responses.activate
    def test_submit_and_generate_manifest_json(self):
        specimen = SubmissionsSpecimen()
        specimen.specimen_id = "specimen9876"
        specimen.biosample_accession = "SAMEA12345678"
        db.session.add(specimen)
        db.session.commit()

        mock_response_from_ena = {"taxId": "6344",
                                  "scientificName": "Arenicola marina",
                                  "commonName": "lugworm",
                                  "formalName": "true",
                                  "rank": "species",
                                  "division": "INV",
                                  "lineage": "",
                                  "geneticCode": "1",
                                  "mitochondrialGeneticCode": "5",
                                  "submittable": "true"}
        responses.add(responses.GET, 'https://www.ebi.ac.uk/ena/taxonomy/rest/tax-id/6344',
                      json=mock_response_from_ena, status=200)
        mock_response_from_tolid = [{
            "species": {
                "commonName": "lugworm",
                "family": "Arenicolidae",
                "genus": "Arenicola",
                "kingdom": "Metazoa",
                "order": "Capitellida",
                "phylum": "Annelida",
                "prefix": "wuAreMari",
                "scientificName": "Arenicola marina",
                "taxaClass": "Polychaeta",
                "taxonomyId": 6344
            },
            "specimen": {
                "specimenId": "specimen9876"
            },
            "tolId": "wuAreMari1"
        }]
        responses.add(responses.POST, os.environ['TOLID_URL'] + '/tol-ids',
                      json=mock_response_from_tolid, status=200)

        mock_response_from_ena = '<?xml version="1.0" encoding="UTF-8"?><?xml-stylesheet type="text/xsl" href="receipt.xsl"?><RECEIPT receiptDate="2021-04-07T12:47:39.998+01:00" submissionFile="tmphz9luulrsubmission_3.xml" success="true"><SAMPLE accession="ERS6206028" alias="' + str(1) + '" status="PRIVATE"><EXT_ID accession="SAMEA8521239" type="biosample"/></SAMPLE><SUBMISSION accession="ERA3819349" alias="SUBMISSION-07-04-2021-12:47:36:825"/><ACTIONS>ADD</ACTIONS></RECEIPT>'  # noqa
        responses.add(responses.POST, os.environ['ENA_URL'] + '/ena/submit/drop-box/submit/',
                      body=mock_response_from_ena, status=200)

        body = {'samples': [
                    {'row': 1,
                     'SPECIMEN_ID': 'specimen9876',
                     'TAXON_ID': 6344,
                     'SCIENTIFIC_NAME': 'Arenicola marina',
                     'GENUS': 'Arenicola',
                     'FAMILY': 'Arenicolidae',
                     'ORDER_OR_GROUP': 'None',
                     'COMMON_NAME': 'lugworm',
                     'LIFESTAGE': 'ADULT',
                     'SEX': 'FEMALE',
                     'ORGANISM_PART': 'MUSCLE',
                     'GAL': 'SANGER INSTITUTE',
                     'GAL_SAMPLE_ID': 'SAN000100',
                     'COLLECTED_BY': 'ALEX COLLECTOR',
                     'COLLECTOR_AFFILIATION': 'THE COLLECTOR INSTITUTE',
                     'DATE_OF_COLLECTION': '2020-09-01',
                     'COLLECTION_LOCATION': 'UNITED KINGDOM | DARK FOREST',
                     'DECIMAL_LATITUDE': '+50.12345678',
                     'DECIMAL_LONGITUDE': '-1.98765432',
                     'HABITAT': 'Woodland',
                     'IDENTIFIED_BY': 'JO IDENTIFIER',
                     'IDENTIFIER_AFFILIATION': 'THE IDENTIFIER INSTITUTE',
                     'VOUCHER_ID': 'voucher1'}
                ]}

        # No authorisation token given
        response = self.client.open(
            '/api/v1/manifests/generate',
            method='POST',
            json=body)
        self.assert401(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Invalid authorisation token given
        response = self.client.open(
            '/api/v1/manifests/generate',
            method='POST',
            headers={"api-key": "12345678"},
            json=body)
        self.assert401(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Not a submitter
        response = self.client.open(
            '/api/v1/manifests/generate',
            method='POST',
            headers={"api-key": self.user1.api_key},
            json=body)
        self.assert403(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Corect - should validate and return errors
        response = self.client.open(
            '/api/v1/manifests/generate',
            method='POST',
            headers={"api-key": self.user3.api_key},
            json=body)
        self.assert200(response,
                       'Response body is : ' + response.data.decode('utf-8'))
        expected = {'manifestId': 1,
                    'submissionStatus': True,
                    'projectName': 'ToL',
                    'stsManifestId': None,
                    'samples': [{
                        'row': 1,
                        'SPECIMEN_ID': 'specimen9876',
                        'TAXON_ID': 6344,
                        'SCIENTIFIC_NAME': 'Arenicola marina',
                        'GENUS': 'Arenicola',
                        'FAMILY': 'Arenicolidae',
                        'ORDER_OR_GROUP': 'None',
                        'COMMON_NAME': 'lugworm',
                        'LIFESTAGE': 'ADULT',
                        'SEX': 'FEMALE',
                        'ORGANISM_PART': 'MUSCLE',
                        'GAL': 'SANGER INSTITUTE',
                        'GAL_SAMPLE_ID': 'SAN000100',
                        'COLLECTED_BY': 'ALEX COLLECTOR',
                        'COLLECTOR_AFFILIATION': 'THE COLLECTOR INSTITUTE',
                        'DATE_OF_COLLECTION': '2020-09-01',
                        'COLLECTION_LOCATION': 'UNITED KINGDOM | DARK FOREST',
                        'DECIMAL_LATITUDE': '+50.12345678',
                        'DECIMAL_LONGITUDE': '-1.98765432',
                        'HABITAT': 'Woodland',
                        'IDENTIFIED_BY': 'JO IDENTIFIER',
                        'IDENTIFIER_AFFILIATION': 'THE IDENTIFIER INSTITUTE',
                        'VOUCHER_ID': 'voucher1',
                        'DEPTH': None,
                        'ELEVATION': None,
                        'RELATIONSHIP': None,
                        'SYMBIONT': None,
                        'CULTURE_OR_STRAIN_ID': None,
                        'tolId': 'wuAreMari1',
                        'biosampleAccession': "SAMEA8521239",
                        'sraAccession': "ERS6206028",
                        'submissionAccession': "ERA3819349",
                        'submissionError': None,
                        'sampleSameAs': None,
                        'sampleDerivedFrom': 'SAMEA12345678',
                        'sampleSymbiontOf': None}
                    ]}
        self.assertEqual(expected, response.json)

    def test_upload_manifest_excel(self):
        # No authorisation token given
        body = []
        response = self.client.open(
            '/api/v1/manifests/upload-excel',
            method='POST',
            json=body)
        self.assert401(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Invalid authorisation token given
        body = []
        response = self.client.open(
            '/api/v1/manifests/upload-excel',
            method='POST',
            headers={"api-key": "12345678"},
            json=body)
        self.assert401(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Excel file missing
        data = {}
        response = self.client.open(
            '/api/v1/manifests/upload-excel',
            method='POST',
            headers={"api-key": self.user3.api_key},
            data=data)
        self.assert400(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Excel file with no taxon ID, specimen ID
        file = open('swagger_server/test/test-manifest-field-missing.xlsx', 'rb')
        data = {
            'excelFile': (file, 'test_file.xlsx'),
        }
        expected = {'manifestId': None,
                    'number_of_errors': 1,
                    'validations': [{'results': [], 'row': 1},
                                    {'results': [{'field': 'SCIENTIFIC_NAME',
                                                  'message': 'A value must be given',
                                                  'severity': 'ERROR'}],
                                     'row': 2}]}
        response = self.client.open(
            '/api/v1/manifests/upload-excel',
            method='POST',
            headers={"api-key": self.user3.api_key},
            data=data)
        file.close()
        self.assert400(response,
                       'Response body is : ' + response.data.decode('utf-8'))
        self.assertEqual(expected, response.json)

        # User not a submitter
        file = open('swagger_server/test/test-manifest.xlsx', 'rb')
        data = {
            'excelFile': (file, 'test_file.xlsx'),
        }
        response = self.client.open(
            '/api/v1/manifests/upload-excel',
            method='POST',
            headers={"api-key": self.user1.api_key},
            data=data)
        file.close()
        self.assert403(response, 'Not received a 403 response')

        # Excel file correct
        file = open('swagger_server/test/test-manifest.xlsx', 'rb')
        data = {
            'excelFile': (file, 'test_file.xlsx'),
        }
        response = self.client.open(
            '/api/v1/manifests/upload-excel',
            method='POST',
            headers={"api-key": self.user3.api_key},
            data=data)
        file.close()
        self.assert200(response, 'Response body is : ' + response.data.decode('utf-8'))
        expected = {'manifestId': 1,
                    'projectName': 'ToL',
                    'stsManifestId': None,
                    'samples': [{
                        'COLLECTED_BY': 'FABIAN HERDER',
                        'COLLECTION_LOCATION': 'Germany | Bonn | Zoological Research Museum Alexander Koenig Leibniz Institute for Animal Biodiversity',  # noqa
                        'COLLECTOR_AFFILIATION': 'ZFMK Adenauerallee 160 53113 Bonn Germany',
                        'COMMON_NAME': None,
                        'CULTURE_OR_STRAIN_ID': None,
                        'DATE_OF_COLLECTION': '2020-12-08',
                        'DECIMAL_LATITUDE': '50.7223215315783',
                        'DECIMAL_LONGITUDE': '7.11382096910185',
                        'DEPTH': None,
                        'ELEVATION': None,
                        'FAMILY': 'Telmatherinidae',
                        'GAL': 'SANGER INSTITUTE',
                        'GAL_SAMPLE_ID': 'SAN2300000',
                        'GENUS': 'Telmatherina',
                        'HABITAT': 'Aquarium ZFMK Bonn',
                        'IDENTIFIED_BY': 'FABIAN HERDER',
                        'IDENTIFIER_AFFILIATION': 'ZFMK Adenauerallee 160 53113 Bonn Germany',
                        'LIFESTAGE': 'ADULT',
                        'ORDER_OR_GROUP': 'Atheriniformes',
                        'ORGANISM_PART': 'TESTIS',
                        'RELATIONSHIP': None,
                        'SCIENTIFIC_NAME': 'Telmatherina bonti',
                        'SEX': 'MALE',
                        'SPECIMEN_ID': 'SAN2300000',
                        'SYMBIONT': None,
                        'TAXON_ID': 446457,
                        'VOUCHER_ID': 'NOT_PROVIDED',
                        'biosampleAccession': None,
                        'row': 1,
                        'sampleDerivedFrom': None,
                        'sampleSameAs': None,
                        'sampleSymbiontOf': None,
                        'sraAccession': None,
                        'submissionAccession': None,
                        'submissionError': None,
                        'tolId': None}, {
                        'COLLECTED_BY': 'FABIAN HERDER',
                        'COLLECTION_LOCATION': 'Germany | Bonn | Zoological Research Museum Alexander Koenig Leibniz Institute for Animal Biodiversity',  # noqa
                        'COLLECTOR_AFFILIATION': 'ZFMK Adenauerallee 160 53113 Bonn Germany',
                        'COMMON_NAME': None,
                        'CULTURE_OR_STRAIN_ID': None,
                        'DATE_OF_COLLECTION': '2020-12-08',
                        'DECIMAL_LATITUDE': '50.7223215315783',
                        'DECIMAL_LONGITUDE': '7.11382096910185',
                        'DEPTH': None,
                        'ELEVATION': None,
                        'FAMILY': 'Telmatherinidae',
                        'GAL': 'SANGER INSTITUTE',
                        'GAL_SAMPLE_ID': 'SAN2300000',
                        'GENUS': 'Telmatherina',
                        'HABITAT': 'Aquarium ZFMK Bonn',
                        'IDENTIFIED_BY': 'FABIAN HERDER',
                        'IDENTIFIER_AFFILIATION': 'ZFMK Adenauerallee 160 53113 Bonn Germany',
                        'LIFESTAGE': 'ADULT',
                        'ORDER_OR_GROUP': 'Atheriniformes',
                        'ORGANISM_PART': 'LIVER',
                        'RELATIONSHIP': None,
                        'SCIENTIFIC_NAME': 'Telmatherina bonti',
                        'SEX': 'MALE',
                        'SPECIMEN_ID': 'SAN2300000',
                        'SYMBIONT': None,
                        'TAXON_ID': 446457,
                        'VOUCHER_ID': 'NOT_PROVIDED',
                        'biosampleAccession': None,
                        'row': 2,
                        'sampleDerivedFrom': None,
                        'sampleSameAs': None,
                        'sampleSymbiontOf': None,
                        'sraAccession': None,
                        'submissionAccession': None,
                        'submissionError': None,
                        'tolId': None}],
                    'submissionStatus': None}
        self.assertEqual(expected, response.json)
        # Has the Excel file been uploaded?
        manifest = db.session.query(SubmissionsManifest) \
            .filter(SubmissionsManifest.manifest_id == 1) \
            .one_or_none()
        self.assertEqual("1.xlsx", manifest.excel_file)

        # Get that from minio and check it is the same file as went in
        dir = tempfile.TemporaryDirectory()
        excel_utils.load_excel(manifest=manifest, dirname=dir.name, filename="test.xlsx")
        self.assertTrue(filecmp.cmp(dir.name + "/test.xlsx",
                                    'swagger_server/test/test-manifest.xlsx',
                                    shallow=False))

    def test_download_manifest_excel(self):
        # Upload it first
        file = open('swagger_server/test/test-manifest.xlsx', 'rb')
        data = {
            'excelFile': (file, 'test_file.xlsx'),
        }
        response = self.client.open(
            '/api/v1/manifests/upload-excel',
            method='POST',
            headers={"api-key": self.user3.api_key},
            data=data)
        file.close()
        self.assert200(response, 'Response body is : ' + response.data.decode('utf-8'))
        # Don't need to test the response here as it has been tested before

        # Pretend all the ID columns have been generated
        manifest = db.session.query(SubmissionsManifest) \
            .filter(SubmissionsManifest.manifest_id == 1) \
            .one_or_none()
        counter = 1
        for sample in manifest.samples:
            for field in SubmissionsSample.id_fields:
                setattr(sample, field["python_name"], str(counter))
                counter += 1
        # Make sure there is a None in there
        manifest.samples[0].sample_symbiont_of = None

        # No authorisation token given
        response = self.client.open(
            '/api/v1/manifests/1/download-excel',
            method='GET')
        self.assert401(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Invalid authorisation token given
        response = self.client.open(
            '/api/v1/manifests/1/download-excel',
            method='GET',
            headers={"api-key": "12345678"})
        self.assert401(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Manifest missing
        response = self.client.open(
            '/api/v1/manifests/999/download-excel',
            method='GET',
            headers={"api-key": self.user3.api_key})
        self.assert404(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # User not a submitter
        response = self.client.open(
            '/api/v1/manifests/1/download-excel',
            method='GET',
            headers={"api-key": self.user1.api_key})
        file.close()
        self.assert403(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        response = self.client.open(
            '/api/v1/manifests/1/download-excel',
            method='GET',
            headers={"api-key": self.user3.api_key})
        file.close()
        self.assert200(response,
                       'Did not receive a 200 response')
        self.assertEqual('application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         response.content_type)

        # Save as Excel file
        file = open('swagger_server/test/test-manifest-validated.xlsx', 'wb')
        file.write(response.get_data())
        file.close()
        workbook = load_workbook(filename='swagger_server/test/test-manifest-validated.xlsx')
        sheet = workbook.active
        # Do we have the columns filled in that we expect
        for sample in manifest.samples:
            for field in SubmissionsSample.id_fields:
                column = excel_utils.find_column(sheet, field["field_name"]) + 1
                row = sample.row + 1
                value = getattr(sample, field["python_name"])
                self.assertEqual(sheet.cell(row=row, column=column).value, value)

    def test_download_manifest_excel_originally_json(self):
        body = {'samples': [
                    {'row': 1,
                     'SPECIMEN_ID': 'specimen9876',
                     'TAXON_ID': 6344,
                     'SCIENTIFIC_NAME': 'Arenicola marina',
                     'GENUS': 'Arenicola',
                     'FAMILY': 'Arenicolidae',
                     'ORDER_OR_GROUP': 'None',
                     'COMMON_NAME': 'lugworm',
                     'LIFESTAGE': 'ADULT',
                     'SEX': 'FEMALE',
                     'ORGANISM_PART': 'MUSCLE',
                     'GAL': 'SANGER INSTITUTE',
                     'GAL_SAMPLE_ID': 'SAN000100',
                     'COLLECTED_BY': 'ALEX COLLECTOR',
                     'COLLECTOR_AFFILIATION': 'THE COLLECTOR INSTITUTE',
                     'DATE_OF_COLLECTION': '2020-09-01',
                     'COLLECTION_LOCATION': 'UNITED KINGDOM | DARK FOREST',
                     'DECIMAL_LATITUDE': '+50.12345678',
                     'DECIMAL_LONGITUDE': '-1.98765432',
                     'HABITAT': 'Woodland',
                     'IDENTIFIED_BY': 'JO IDENTIFIER',
                     'IDENTIFIER_AFFILIATION': 'THE IDENTIFIER INSTITUTE',
                     'VOUCHER_ID': 'voucher1'}
                ]}

        # Correct, full JSON
        response = self.client.open(
            '/api/v1/manifests',
            method='POST',
            headers={"api-key": self.user3.api_key},
            json=body)
        self.assert200(response,
                       'Response body is : ' + response.data.decode('utf-8'))
        # Don't need to test the response here as it has been tested before

        # Pretend all the ID columns have been generated
        manifest = db.session.query(SubmissionsManifest) \
            .filter(SubmissionsManifest.manifest_id == 1) \
            .one_or_none()
        counter = 1
        for sample in manifest.samples:
            for field in SubmissionsSample.id_fields:
                setattr(sample, field["python_name"], str(counter))
                counter += 1
        # Make sure there is a None in there
        manifest.samples[0].sample_symbiont_of = None

        response = self.client.open(
            '/api/v1/manifests/1/download-excel',
            method='GET',
            headers={"api-key": self.user3.api_key})
        self.assert200(response,
                       'Did not receive a 200 response')
        self.assertEqual('application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         response.content_type)

        # Save as Excel file
        file = open('swagger_server/test/test-manifest-validated.xlsx', 'wb')
        file.write(response.get_data())
        file.close()
        workbook = load_workbook(filename='swagger_server/test/test-manifest-validated.xlsx')
        sheet = workbook.active
        # Do we have the columns filled in that we expect
        for sample in manifest.samples:
            for field in SubmissionsSample.id_fields:
                column = excel_utils.find_column(sheet, field["field_name"]) + 1
                row = sample.row + 1
                value = getattr(sample, field["python_name"])
                self.assertEqual(sheet.cell(row=row, column=column).value, value)

        # Do we have the original columns in?
        for sample in manifest.samples:
            for field in SubmissionsSample.all_fields:
                column = excel_utils.find_column(sheet, field["field_name"]) + 1
                row = sample.row + 1
                value = getattr(sample, field["python_name"])
                self.assertEqual(sheet.cell(row=row, column=column).value, value)


if __name__ == '__main__':
    import unittest
    unittest.main()
