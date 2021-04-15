from openpyxl import load_workbook
import re
import os
import json
import shutil
from swagger_server.model import SubmissionsManifest, SubmissionsSample
from minio import Minio
from minio.error import ResponseError, BucketAlreadyOwnedByYou, BucketAlreadyExists


def clean_cell(value):
    if value is None:
        return None
    value = str(value)
    value = re.sub(r"\s+", ' ', value)
    value = re.sub(r"\s+$", '', value)
    value = re.sub(r"^\s+", '', value)
    return value


def get_columns(sheet):
    columns = {}
    row = sheet[1]
    for cell in row:
        if cell.value != "":
            columns[cell.value] = cell.column - 1
    return columns


def find_column(sheet, column_heading):
    columns = get_columns(sheet)
    if column_heading in columns:
        return columns[column_heading]

    return None  # Couldn't find the column


def cell_value(sheet, row, column_heading):
    column_index = find_column(sheet, column_heading)
    if column_index is None:
        return ""
    return clean_cell(row[column_index])


def read_excel(dirname=None, filename=None,
               user=None, project_name=None):
    workbook = load_workbook(filename=dirname+'/'+filename)
    sheet = workbook.active

    manifest = SubmissionsManifest()
    if project_name is not None:
        manifest.project_name = project_name
    manifest.user = user

    current_row = 2
    for row in sheet.iter_rows(min_row=current_row,
                               max_row=sheet.max_row, values_only=True):
        # Only read in rows that have a value in the SERIES column
        if cell_value(sheet, row, "SERIES") is not None:
            sample = SubmissionsSample()
            sample.manifest = manifest

            sample.row = current_row - 1
            sample.specimen_id = cell_value(sheet, row, "SPECIMEN_ID")
            sample.taxonomy_id = int(cell_value(sheet, row, "TAXON_ID"))
            sample.scientific_name = cell_value(sheet, row, "SCIENTIFIC_NAME")
            sample.family = cell_value(sheet, row, "FAMILY")
            sample.genus = cell_value(sheet, row, "GENUS")
            sample.order_or_group = cell_value(sheet, row, "ORDER_OR_GROUP")
            sample.common_name = cell_value(sheet, row, "COMMON_NAME")
            sample.lifestage = cell_value(sheet, row, "LIFESTAGE")
            sample.sex = cell_value(sheet, row, "SEX")
            sample.organism_part = cell_value(sheet, row, "ORGANISM_PART")
            sample.GAL = cell_value(sheet, row, "GAL")
            sample.GAL_sample_id = cell_value(sheet, row, "GAL_SAMPLE_ID")
            sample.collected_by = cell_value(sheet, row, "COLLECTED_BY")
            sample.collector_affiliation = cell_value(sheet, row, "COLLECTOR_AFFILIATION")
            date_of_collection = cell_value(sheet, row, "DATE_OF_COLLECTION")
            if date_of_collection is not None:
                # Only the date part
                sample.date_of_collection = date_of_collection.partition(' ')[0]
            sample.collection_location = cell_value(sheet, row, "COLLECTION_LOCATION")
            sample.decimal_latitude = cell_value(sheet, row, "DECIMAL_LATITUDE")
            sample.decimal_longitude = cell_value(sheet, row, "DECIMAL_LONGITUDE")
            sample.habitat = cell_value(sheet, row, "HABITAT")
            sample.identified_by = cell_value(sheet, row, "IDENTIFIED_BY")
            sample.identifier_affiliation = cell_value(sheet, row, "IDENTIFIER_AFFILIATION")
            sample.voucher_id = cell_value(sheet, row, "VOUCHER_ID")
            sample.elevation = cell_value(sheet, row, "ELEVATION")
            sample.depth = cell_value(sheet, row, "DEPTH")
            sample.relationship = cell_value(sheet, row, "RELATIONSHIP")
            sample.symbiont = cell_value(sheet, row, "SYMBIONT")
            sample.culture_or_strain_id = cell_value(sheet, row, "CULTURE_OR_STRAIN_ID")

        current_row += 1
    return manifest


def save_excel(manifest=None, dirname=None, filename=None):
    minio_client = Minio(os.getenv("MINIO_URI"),
                         os.getenv("MINIO_ACCESS_KEY"),
                         os.getenv("MINIO_SECRET_KEY"),
                         secure=(os.getenv("MINIO_SECURE", False) == "True"))
    bucket = os.getenv("MINIO_BUCKET")
    policy_read_only = {
        "Version": "2012-10-17",
        "Statement": [
            {
                "Sid": "",
                "Effect": "Allow",
                "Principal": {"AWS": "*"},
                "Action": "s3:GetBucketLocation",
                "Resource": "arn:aws:s3:::" + bucket
            },
            {
                "Sid": "",
                "Effect": "Allow",
                "Principal": {"AWS": "*"},
                "Action": "s3:ListBucket",
                "Resource": "arn:aws:s3:::" + bucket
            },
            {
                "Sid": "",
                "Effect": "Allow",
                "Principal": {"AWS": "*"},
                "Action": "s3:GetObject",
                "Resource": "arn:aws:s3:::" + bucket + "/*"
            }
        ]
    }

    try:
        minio_client.make_bucket(bucket)
        minio_client.set_bucket_policy(bucket, json.dumps(policy_read_only))
    except BucketAlreadyOwnedByYou:
        pass
    except BucketAlreadyExists:
        pass
    except ResponseError:
        pass

    try:
        (root, ext) = os.path.splitext(filename)
        minio_filename = str(manifest.manifest_id) + ext

        minio_client.fput_object(
            bucket, minio_filename, os.path.join(dirname, filename)
        )
        manifest.excel_file = minio_filename
    except ResponseError:
        pass
    # Needs committing elsewhere


def load_excel(manifest=None, dirname=None, filename=None):
    if manifest.excel_file is not None:
        minio_client = Minio(os.getenv("MINIO_URI"),
                             os.getenv("MINIO_ACCESS_KEY"),
                             os.getenv("MINIO_SECRET_KEY"),
                             secure=(os.getenv("MINIO_SECURE", False) == "True"))
        try:
            minio_client.fget_object(
                os.getenv("MINIO_BUCKET"), manifest.excel_file, os.path.join(dirname, filename),
            )
            return True
        except ResponseError:
            return False
    return False


def create_excel(manifest=None, dirname=None, filename=None):
    shutil.copyfile("swagger_server/templates/manifest_empty.xlsx",
                    os.path.join(dirname, filename))
    add_columns(manifest=manifest, fields=SubmissionsSample.all_fields,
                dirname=dirname, filename=filename)


def add_columns(manifest=None, fields=None, dirname=None, filename=None):
    workbook = load_workbook(filename=os.path.join(dirname, filename))
    sheet = workbook.active

    # Sort out column headings
    columns = get_columns(sheet)
    for field in fields:
        column_heading = field["field_name"]
        if column_heading not in columns:
            columns[column_heading] = len(columns) - 1
            sheet.cell(row=1, column=columns[column_heading] + 1,
                       value=column_heading)

    # Now the sheet has all column headings and find_column will find them
    for sample in manifest.samples:
        for field in fields:
            column = find_column(sheet, field["field_name"]) + 1  # add 1 to the index
            row = sample.row + 1  # sample.row is the index of the row
            value = getattr(sample, field["python_name"])
            sheet.cell(column=column, row=row, value=value)

    workbook.save(os.path.join(dirname, filename))
