import datetime
import json
import subprocess

import jsonpath_rw_ext as jp
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from COPO.web.apps.web_copo.lookup import lookup as lk
from COPO.web.apps.web_copo.schemas.utils.data_utils import json_to_pytype


def make_tax_from_sample(s):
    out = dict()
    out["SYMBIONT"] = "symbiont"
    out["TAXON_ID"] = s["TAXON_ID"]
    out["ORDER_OR_GROUP"] = s["ORDER_OR_GROUP"]
    out["FAMILY"] = s["FAMILY"]
    out["GENUS"] = s["GENUS"]
    out["SCIENTIFIC_NAME"] = s["SCIENTIFIC_NAME"]
    out["INFRASPECIFIC_EPITHET"] = s["INFRASPECIFIC_EPITHET"]
    out["CULTURE_OR_STRAIN_ID"] = s["CULTURE_OR_STRAIN_ID"]
    out["COMMON_NAME"] = s["COMMON_NAME"]
    out["TAXON_REMARKS"] = s["TAXON_REMARKS"]
    out["RACK_OR_PLATE_ID"] = s["RACK_OR_PLATE_ID"]
    out["TUBE_OR_WELL_ID"] = s["TUBE_OR_WELL_ID"]
    for el in out:
        out[el] = str(out[el]).strip()
    return out


def validate_date(date_text):
    try:
        datetime.datetime.strptime(date_text, '%Y-%m-%d')
    except ValueError:
        raise ValueError("Incorrect data format, should be YYYY-MM-DD")


def check_taxon_ena_submittable(taxon):
    errors = []
    receipt = None
    taxinfo = None
    curl_cmd = "curl " + "https://www.ebi.ac.uk/ena/taxonomy/rest/tax-id/" + taxon
    try:
        receipt = subprocess.check_output(curl_cmd, shell=True)
        print(receipt)
        taxinfo = json.loads(receipt.decode("utf-8"))
        if taxinfo["submittable"] != 'true':
            errors.append("TAXON_ID " + taxon + " is not submittable to ENA")
    except Exception as e:
        if receipt:
            try:
                errors.append(
                    "ENA returned - " + taxinfo.get("error", "no error returned") + " - for TAXON_ID " + taxon)
            except NameError:
                errors.append(
                    "ENA returned - " + receipt.decode("utf-8") + " - for TAXON_ID " + taxon)
        else:
            errors.append("No response from ENA taxonomy for taxon " + taxon)
    return errors


def create_barcoding_spreadsheet():
    # get barcoding fields
    s = json_to_pytype(lk.WIZARD_FILES["sample_details"])

    # to get all elements with version starting with bc
    # fields = jp.match('$.properties[?(@.specifications[*]~".*bc.*" & @.required=="true")].versions[0]', s)

    barcode_fields = jp.match(
        '$.properties[?(@.specifications[*] == "bc" & @.required=="true")].versions['
        '0]',
        s)
    amplicon_fields = jp.match(
        '$.properties[?(@.specifications[*] == "bc_amp" & @.required=="true")].versions['
        '0]',
        s)

    wb = Workbook()
    sheet_id = wb.get_sheet_by_name("Sheet")
    wb.remove_sheet(sheet_id)
    wb.create_sheet("Specimens")
    wb.create_sheet("Amplicons")

    spec = wb.worksheets[0]
    for idx, f in enumerate(barcode_fields):
        spec.cell(row=1, column=idx + 1).value = f

    amp = wb.worksheets[1]
    for idx, f in enumerate(amplicon_fields):
        amp.cell(row=1, column=idx + 1).value = f

    for sheet in wb.worksheets:
        for idx, column in enumerate(sheet.columns):
            cell = sheet.cell(column=idx + 1, row=1)
            sheet.column_dimensions[get_column_letter(idx + 1)].width = len(cell.value) + 5

    return wb
